"""
Step 4 Duplicate Remover Module
Removes duplicate rows from Step 3 files based on columns A-K while preserving unique entries.
"""

import openpyxl
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple
import logging

class Step4DuplicateRemover:
    """
    Processes Step 3 files by removing duplicate rows based on columns A-K
    """
    
    def __init__(self, output_dir: Optional[str] = None):
        """
        Initialize the Step 4 duplicate remover
        
        Args:
            output_dir: Directory to save Step 4 files (default: ./output)
        """
        if output_dir:
            self.output_dir = Path(output_dir)
        else:
            self.output_dir = Path("output")
        
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Columns A through K (1-11) are used for duplicate detection
        self.comparison_columns = list(range(1, 12))  # Columns 1-11 (A-K)
    
    def extract_row_data(self, worksheet, row_num: int) -> Tuple[str, List[str]]:
        """
        Extract data from columns A-K for duplicate comparison
        
        Args:
            worksheet: openpyxl worksheet object
            row_num: Row number to extract data from
            
        Returns:
            Tuple of (comparison_key, full_row_data)
        """
        try:
            # Extract data from columns A-K for comparison
            comparison_data = []
            full_row_data = []
            
            # Get data from all columns up to the last column with data
            max_col = min(worksheet.max_column, 20)  # Limit to first 20 columns for efficiency
            
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row_num, col).value
                cell_str = str(cell_value).strip() if cell_value is not None else ""
                
                full_row_data.append(cell_str)
                
                # Only use columns A-K for comparison
                if col in self.comparison_columns:
                    comparison_data.append(cell_str)
            
            # Create comparison key from columns A-K
            comparison_key = "|".join(comparison_data)
            
            return comparison_key, full_row_data
            
        except Exception as e:
            logging.error(f"Error extracting row data from row {row_num}: {str(e)}")
            return "", []
    
    def remove_duplicates_from_step3(self, step3_file_path: str) -> str:
        """
        Remove duplicate rows from Step 3 file and save as Step 4
        
        Args:
            step3_file_path: Path to Step 3 file
            
        Returns:
            Path to created Step 4 file
        """
        try:
            # Load Step 3 file
            wb = openpyxl.load_workbook(step3_file_path)
            ws = wb.active
            
            logging.info(f"Processing Step 3 file: {Path(step3_file_path).name}")
            logging.info(f"Original file has {ws.max_row} rows")
            
            # Track unique rows using comparison key
            seen_keys = set()
            unique_rows = []
            duplicate_count = 0
            
            # Process data rows starting from row 4
            data_start_row = 4
            
            for row_num in range(data_start_row, ws.max_row + 1):
                comparison_key, full_row_data = self.extract_row_data(ws, row_num)
                
                # Skip empty rows
                if not any(cell.strip() for cell in full_row_data if cell):
                    continue
                
                # Check if this combination of A-K columns has been seen before
                if comparison_key not in seen_keys and comparison_key.strip():
                    seen_keys.add(comparison_key)
                    unique_rows.append((row_num, full_row_data))
                    logging.debug(f"Row {row_num}: UNIQUE - {comparison_key[:100]}...")
                else:
                    duplicate_count += 1
                    logging.debug(f"Row {row_num}: DUPLICATE - {comparison_key[:100]}...")
            
            logging.info(f"Found {len(unique_rows)} unique rows, removed {duplicate_count} duplicates")
            
            # Create new workbook with unique rows
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            
            # Copy header rows (rows 1-3) from original
            for row_num in range(1, min(data_start_row, ws.max_row + 1)):
                for col in range(1, ws.max_column + 1):
                    source_cell = ws.cell(row_num, col)
                    target_cell = new_ws.cell(row_num, col)
                    
                    # Copy value
                    target_cell.value = source_cell.value
                    
                    # Copy formatting
                    if source_cell.has_style:
                        target_cell.font = source_cell.font.copy()
                        target_cell.border = source_cell.border.copy()
                        target_cell.fill = source_cell.fill.copy()
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = source_cell.protection.copy()
                        target_cell.alignment = source_cell.alignment.copy()
            
            # Copy column dimensions (widths) from original worksheet
            for col_letter, col_dim in ws.column_dimensions.items():
                new_ws.column_dimensions[col_letter].width = col_dim.width
            
            # Copy row dimensions (heights) for header rows
            for row_num in range(1, min(data_start_row, ws.max_row + 1)):
                if row_num in ws.row_dimensions:
                    new_ws.row_dimensions[row_num].height = ws.row_dimensions[row_num].height
            
            # Write unique data rows
            current_row = data_start_row
            for original_row_num, row_data in unique_rows:
                # Copy row height from original if it exists
                if original_row_num in ws.row_dimensions:
                    new_ws.row_dimensions[current_row].height = ws.row_dimensions[original_row_num].height
                
                for col_idx, cell_value in enumerate(row_data, 1):
                    if cell_value:  # Only write non-empty values
                        new_cell = new_ws.cell(current_row, col_idx)
                        new_cell.value = cell_value
                        
                        # Copy formatting from original cell
                        source_cell = ws.cell(original_row_num, col_idx)
                        if source_cell.has_style:
                            new_cell.font = source_cell.font.copy()
                            new_cell.border = source_cell.border.copy()
                            new_cell.fill = source_cell.fill.copy()
                            new_cell.number_format = source_cell.number_format
                            new_cell.protection = source_cell.protection.copy()
                            new_cell.alignment = source_cell.alignment.copy()
                
                current_row += 1
            
            # Generate Step 4 filename
            step3_path = Path(step3_file_path)
            step4_filename = step3_path.name.replace("Step3.xlsx", "Step4.xlsx")
            step4_path = self.output_dir / step4_filename
            
            # Save Step 4 file
            new_wb.save(str(step4_path))
            logging.info(f"Created Step 4 file: {step4_path} with {len(unique_rows)} unique rows")
            
            return str(step4_path)
            
        except Exception as e:
            logging.error(f"Error removing duplicates from Step 3 file: {str(e)}")
            raise Exception(f"Failed to create Step 4 file: {str(e)}")
    
    def process_sheet_to_step4(self, step3_file_path: str) -> Optional[str]:
        """
        Complete Step 4 processing for a single Step 3 file
        
        Args:
            step3_file_path: Path to the Step 3 file
            
        Returns:
            Path to created Step 4 file, or None if failed
        """
        try:
            step4_file_path = self.remove_duplicates_from_step3(step3_file_path)
            return step4_file_path
            
        except Exception as e:
            logging.error(f"Error processing Step 3 file to Step 4: {str(e)}")
            return None
    
    def process_multiple_sheets_to_step4(self, step3_files: list) -> Dict[str, Any]:
        """
        Process multiple Step 3 files to Step 4
        
        Args:
            step3_files: List of dictionaries with Step 3 file information
            
        Returns:
            Dictionary with processing results
        """
        results = {
            'success_count': 0,
            'failed_count': 0,
            'step4_files': [],
            'failed_sheets': [],
            'total_sheets': len(step3_files)
        }
        
        for step3_info in step3_files:
            sheet_name = step3_info['sheet_name']
            step3_file_path = step3_info['file_path']
            
            try:
                step4_file_path = self.process_sheet_to_step4(step3_file_path)
                
                if step4_file_path:
                    results['success_count'] += 1
                    results['step4_files'].append({
                        'sheet_name': sheet_name,
                        'filename': Path(step4_file_path).name,
                        'file_path': step4_file_path,
                        'step3_file': step3_file_path
                    })
                else:
                    results['failed_count'] += 1
                    results['failed_sheets'].append(sheet_name)
                    
            except Exception as e:
                logging.error(f"Failed to process sheet '{sheet_name}' to Step 4: {str(e)}")
                results['failed_count'] += 1
                results['failed_sheets'].append(sheet_name)
        
        return results