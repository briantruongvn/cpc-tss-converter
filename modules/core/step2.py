"""Step 2 — Article info extraction and template population."""

import logging
import re
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from typing import List, Optional, Tuple

from streamlit.runtime.uploaded_file_manager import UploadedFile

from .config_manager import ConfigManager


class CoreStep2Processor:
    """Core Step 2 processor using unified configuration"""

    def __init__(self, output_dir: Optional[str] = None, config_dir: Optional[str] = None):
        self.config_dir = config_dir
        self.processing_config = ConfigManager.get_processing_config(config_dir)

        if output_dir:
            self.output_dir = Path(output_dir)
        else:
            output_dir_name = self.processing_config.get("general", {}).get("output_dir", "output")
            self.output_dir = Path(output_dir_name)

        self.output_dir.mkdir(parents=True, exist_ok=True)

        step2_config = self.processing_config["step2_config"]
        self.product_name_patterns = step2_config["product_name_patterns"]
        self.article_number_patterns = step2_config["article_number_patterns"]

    def extract_article_info_from_uploaded_file(
        self, uploaded_file: UploadedFile, sheet_name: str
    ) -> Tuple[List[str], List[str]]:
        """Extract article names and numbers from uploaded Streamlit file."""
        try:
            df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)
            return self._search_article_info_in_dataframe(df)
        except Exception as e:
            logging.error(f"Error extracting article info from sheet '{sheet_name}': {str(e)}")
            return [], []

    def extract_article_info_from_file(
        self, file_path: str, sheet_name: str
    ) -> Tuple[List[str], List[str]]:
        """Extract article names and numbers from a local file."""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            return self._search_article_info_in_dataframe(df)
        except Exception as e:
            logging.error(f"Error extracting article info from sheet '{sheet_name}': {str(e)}")
            return [], []

    def _search_article_info_in_dataframe(
        self, df: pd.DataFrame
    ) -> Tuple[List[str], List[str]]:
        """Search for article information in a pandas DataFrame and return arrays."""
        article_name = None
        article_number = None

        # Only search rows 1-12 (0-indexed 0-11)
        for row_idx, row in df.iterrows():
            if row_idx > 11:
                break

            for col_idx, cell_value in enumerate(row):
                if pd.isna(cell_value):
                    continue

                cell_str = str(cell_value).strip()

                if article_name is None:
                    for pattern in self.product_name_patterns:
                        if pattern in cell_str:
                            article_name = self._extract_value_after_pattern(cell_str, pattern)
                            if not article_name and col_idx + 1 < len(row):
                                next_cell = row.iloc[col_idx + 1]
                                if not pd.isna(next_cell):
                                    article_name = str(next_cell).strip()
                            break

                if article_number is None:
                    for pattern in self.article_number_patterns:
                        if pattern in cell_str:
                            article_number = self._extract_value_after_pattern(cell_str, pattern)
                            if not article_number and col_idx + 1 < len(row):
                                next_cell = row.iloc[col_idx + 1]
                                if not pd.isna(next_cell):
                                    article_number = str(next_cell).strip()
                            break

                if article_name and article_number:
                    break

            if article_name and article_number:
                break

        article_names = []
        if article_name:
            names = re.split(r'[\n;]+', article_name.strip())
            article_names = [n.strip() for n in names if n.strip()]

        article_numbers = []
        if article_number:
            numbers = re.split(r'[,\n;]+', article_number.strip())
            article_numbers = [n.strip() for n in numbers if n.strip()]

        return article_names, article_numbers

    def _extract_value_after_pattern(self, text: str, pattern: str) -> Optional[str]:
        """Extract value that comes after a pattern in text."""
        try:
            pattern_index = text.find(pattern)
            if pattern_index == -1:
                return None

            after_pattern = text[pattern_index + len(pattern):].strip()

            while after_pattern and after_pattern[0] in ['::', ':', ' ']:
                after_pattern = after_pattern[1:].strip()

            return after_pattern if after_pattern else None

        except Exception:
            return None

    def populate_template_with_article_info(
        self,
        template_path: str,
        article_names: List[str],
        article_numbers: List[str],
    ) -> str:
        """
        Populate Step 1 template with extracted article information.

        Supports multiple values placed in consecutive columns (R, S, T, …).
        Returns path to the created Step 2 file.
        """
        try:
            step2_path = self._get_step2_filename(template_path)

            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            template_config = ConfigManager.get_template_config()
            layout_config = template_config["layout"]["article_info_rows"]
            article_style = template_config["article_info_style"]

            if article_names:
                article_name_row = layout_config["article_name_row"]
                article_name_merge_end = layout_config["article_name_merge_end"]
                start_column_letter = layout_config["article_name_start_column"]
                start_column_index = openpyxl.utils.column_index_from_string(start_column_letter)

                for i, article_name in enumerate(article_names):
                    current_column_index = start_column_index + i
                    current_column_letter = openpyxl.utils.get_column_letter(current_column_index)

                    merge_range = f"{current_column_letter}{article_name_row}:{current_column_letter}{article_name_merge_end}"
                    ws.merge_cells(merge_range)

                    cell = ws.cell(article_name_row, current_column_index, article_name)
                    logging.info(
                        f"Placing article name '{article_name}' in "
                        f"{current_column_letter}{article_name_row} with "
                        f"rotation {article_style['article_name_alignment']['text_rotation']}°"
                    )

                    cell.fill = PatternFill(
                        start_color=article_style["fill"]["start_color"][2:],
                        end_color=article_style["fill"]["end_color"][2:],
                        fill_type=article_style["fill"]["fill_type"]
                    )
                    cell.font = Font(
                        bold=article_style["font"]["bold"],
                        color=article_style["font"]["color"][2:]
                    )
                    cell.alignment = Alignment(
                        text_rotation=article_style["article_name_alignment"]["text_rotation"],
                        horizontal=article_style["article_name_alignment"]["horizontal"],
                        vertical=article_style["article_name_alignment"]["vertical"],
                        wrap_text=article_style["article_name_alignment"].get("wrap_text", False)
                    )

            if article_numbers:
                article_number_start_row = layout_config["article_number_start_row"]
                start_column_letter = layout_config["article_number_start_column"]
                start_column_index = openpyxl.utils.column_index_from_string(start_column_letter)

                for i, article_number in enumerate(article_numbers):
                    current_column_index = start_column_index + i
                    current_column_letter = openpyxl.utils.get_column_letter(current_column_index)

                    cell = ws.cell(article_number_start_row, current_column_index, article_number)
                    logging.info(
                        f"Placing article number '{article_number}' in "
                        f"{current_column_letter}{article_number_start_row}"
                    )

                    cell.fill = PatternFill(
                        start_color=article_style["fill"]["start_color"][2:],
                        end_color=article_style["fill"]["end_color"][2:],
                        fill_type=article_style["fill"]["fill_type"]
                    )
                    cell.font = Font(
                        bold=article_style["font"]["bold"],
                        color=article_style["font"]["color"][2:]
                    )
                    cell.alignment = Alignment(
                        horizontal=article_style["alignment"]["horizontal"],
                        vertical=article_style["alignment"]["vertical"]
                    )

            wb.save(step2_path)
            logging.info(f"Created Step 2 file: {step2_path}")
            return step2_path

        except Exception as e:
            logging.error(f"Error populating template: {str(e)}")
            raise

    def _add_checkbox_markings(
        self,
        worksheet,
        article_names: List[str],
        article_numbers: List[str],
        layout_config: dict,
    ):
        """Add X marks in data rows for columns that contain article names/numbers."""
        try:
            data_start_row = layout_config.get("data_start_row", 11)
            start_column_letter = layout_config["article_info_rows"]["article_name_start_column"]
            start_column_index = openpyxl.utils.column_index_from_string(start_column_letter)

            num_article_columns = max(len(article_names), len(article_numbers))
            if num_article_columns == 0:
                logging.info("No article info found, skipping checkbox markings")
                return

            logging.info(
                f"Adding X marks to {num_article_columns} article column(s) "
                f"starting from row {data_start_row}"
            )

            markings_added = 0
            for row_num in range(data_start_row, worksheet.max_row + 1):
                row_has_data = any(
                    worksheet.cell(row_num, col).value and
                    str(worksheet.cell(row_num, col).value).strip()
                    for col in range(1, 15)
                )
                if not row_has_data:
                    continue

                for i in range(num_article_columns):
                    worksheet.cell(row_num, start_column_index + i, "X")
                    markings_added += 1

            logging.info(f"Added {markings_added} X marks to article columns")

        except Exception as e:
            logging.error(f"Error adding checkbox markings: {str(e)}")

    def _get_step2_filename(self, step1_path: str) -> str:
        """Generate Step 2 filename from Step 1 path."""
        step1_file = Path(step1_path)
        base_name = step1_file.stem.replace(" - Step1", "")
        filename_template = self.processing_config["general"]["file_naming"]["step2"]

        parts = base_name.split(" - ")
        if len(parts) >= 2:
            original_base = " - ".join(parts[:-1])
            sheet_name = parts[-1]
        else:
            original_base = base_name
            sheet_name = "Unknown"

        step2_filename = filename_template.format(base_name=original_base, sheet_name=sheet_name)
        return str(self.output_dir / step2_filename)
