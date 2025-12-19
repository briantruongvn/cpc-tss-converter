"""
Step 2 Processor Module
Extracts article name and article number from original Excel sheets and populates Step 1 templates.
"""

import openpyxl
import pandas as pd
from pathlib import Path
from typing import Optional, Dict, Any, Tuple
import logging
import re
from streamlit.runtime.uploaded_file_manager import UploadedFile
import tempfile

class Step2Processor:
    """
    Processes Step 1 templates by extracting article information from original Excel sheets
    """
    
    def __init__(self, output_dir: Optional[str] = None):
        """
        Initialize the Step 2 processor
        
        Args:
            output_dir: Directory to save Step 2 files (default: ./output)
        """
        if output_dir:
            self.output_dir = Path(output_dir)
        else:
            self.output_dir = Path("output")
        
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Search patterns for article information
        self.product_name_patterns = [
            "Tên sản phẩm/Product name:",
            "Tên sản phẩm/Product name",
            "Product name:",
            "Tên sản phẩm:",
            "Product name",
            "Tên sản phẩm"
        ]
        
        self.article_number_patterns = [
            "Mã số sản phẩm/Article number:",
            "Mã số sản phẩm/Article number",
            "Article number:",
            "Mã số sản phẩm:",
            "Article number",
            "Mã số sản phẩm"
        ]
    
    def search_text_in_sheet(self, worksheet, search_patterns: list) -> Optional[Tuple[int, int]]:
        """
        Search for text patterns in a worksheet and return cell coordinates
        
        Args:
            worksheet: openpyxl worksheet object
            search_patterns: List of text patterns to search for
            
        Returns:
            Tuple of (row, column) if found, None otherwise
        """
        try:
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row, col).value
                    
                    if cell_value and isinstance(cell_value, str):
                        cell_text = cell_value.strip().lower()
                        
                        for pattern in search_patterns:
                            pattern_lower = pattern.lower().strip()
                            
                            # Exact match or pattern contains the cell text
                            if (pattern_lower == cell_text or 
                                pattern_lower in cell_text or 
                                cell_text in pattern_lower):
                                logging.info(f"Found pattern '{pattern}' at row {row}, col {col}")
                                return (row, col)
            
            return None
            
        except Exception as e:
            logging.error(f"Error searching text in sheet: {str(e)}")
            return None
    
    def extract_adjacent_cell_value(self, worksheet, row: int, col: int) -> str:
        """
        Extract value from the cell to the right of the found pattern
        
        Args:
            worksheet: openpyxl worksheet object
            row: Row number of the found pattern
            col: Column number of the found pattern
            
        Returns:
            String value from adjacent cell, or empty string if not found
        """
        try:
            # Get value from cell to the right (same row, next column)
            adjacent_cell = worksheet.cell(row, col + 1)
            value = adjacent_cell.value
            
            if value is not None:
                # Convert to string and clean up
                clean_value = str(value).strip()
                logging.info(f"Extracted value: '{clean_value}' from row {row}, col {col + 1}")
                return clean_value
            else:
                logging.warning(f"Adjacent cell at row {row}, col {col + 1} is empty")
                return ""
                
        except Exception as e:
            logging.error(f"Error extracting adjacent cell value: {str(e)}")
            return ""
    
    def extract_article_info_from_sheet(self, uploaded_file: UploadedFile, sheet_name: str) -> Dict[str, str]:
        """
        Extract article name and number from a specific sheet in the original file
        
        Args:
            uploaded_file: Streamlit uploaded file object
            sheet_name: Name of the sheet to process
            
        Returns:
            Dictionary with 'product_name' and 'article_number' keys
        """
        try:
            # Save uploaded file temporarily
            with tempfile.NamedTemporaryFile(delete=False, suffix=Path(uploaded_file.name).suffix) as tmp_file:
                tmp_file.write(uploaded_file.getvalue())
                tmp_file.flush()
                
                # Load workbook and get specific sheet
                wb = openpyxl.load_workbook(tmp_file.name, data_only=True)
                
                if sheet_name not in wb.sheetnames:
                    logging.error(f"Sheet '{sheet_name}' not found in workbook")
                    return {'product_name': '', 'article_number': ''}
                
                ws = wb[sheet_name]
                
                # Search for product name
                product_name = ""
                product_name_coords = self.search_text_in_sheet(ws, self.product_name_patterns)
                if product_name_coords:
                    product_name = self.extract_adjacent_cell_value(ws, *product_name_coords)
                
                # Search for article number
                article_number = ""
                article_number_coords = self.search_text_in_sheet(ws, self.article_number_patterns)
                if article_number_coords:
                    article_number = self.extract_adjacent_cell_value(ws, *article_number_coords)
                
                # Clean up temporary file
                Path(tmp_file.name).unlink()
                
                logging.info(f"Extracted from sheet '{sheet_name}': Product='{product_name}', Article='{article_number}'")
                
                return {
                    'product_name': product_name,
                    'article_number': article_number
                }
                
        except Exception as e:
            logging.error(f"Error extracting article info from sheet '{sheet_name}': {str(e)}")
            return {'product_name': '', 'article_number': ''}
    
    def populate_step1_template(self, step1_file_path: str, article_info: Dict[str, str]) -> str:
        """
        Populate Step 1 template with article information and save as Step 2
        
        Args:
            step1_file_path: Path to Step 1 template file
            article_info: Dictionary with product_name and article_number
            
        Returns:
            Path to created Step 2 file
        """
        try:
            # Load Step 1 template
            wb = openpyxl.load_workbook(step1_file_path)
            ws = wb.active
            
            # Populate B1 with product name (if found)
            if article_info['product_name']:
                ws['B1'].value = article_info['product_name']
                logging.info(f"Set B1 to: {article_info['product_name']}")
            
            # Populate B2 with article number (if found)
            if article_info['article_number']:
                ws['B2'].value = article_info['article_number']
                logging.info(f"Set B2 to: {article_info['article_number']}")
            
            # Generate Step 2 filename
            step1_path = Path(step1_file_path)
            step2_filename = step1_path.name.replace("Step1.xlsx", "Step2.xlsx")
            step2_path = self.output_dir / step2_filename
            
            # Save as Step 2 file
            wb.save(str(step2_path))
            logging.info(f"Created Step 2 file: {step2_path}")
            
            return str(step2_path)
            
        except Exception as e:
            logging.error(f"Error populating Step 1 template: {str(e)}")
            raise Exception(f"Failed to create Step 2 file: {str(e)}")
    
    def process_sheet_to_step2(self, uploaded_file: UploadedFile, sheet_name: str, step1_file_path: str) -> Optional[str]:
        """
        Complete Step 2 processing for a single sheet
        
        Args:
            uploaded_file: Original uploaded Excel file
            sheet_name: Name of the sheet to process
            step1_file_path: Path to the Step 1 template file
            
        Returns:
            Path to created Step 2 file, or None if failed
        """
        try:
            # Extract article information from original sheet
            article_info = self.extract_article_info_from_sheet(uploaded_file, sheet_name)
            
            # Check if any information was found
            if not article_info['product_name'] and not article_info['article_number']:
                logging.warning(f"No article information found in sheet '{sheet_name}'")
                # Still create Step 2 file but with empty values
            
            # Populate template and create Step 2 file
            step2_file_path = self.populate_step1_template(step1_file_path, article_info)
            
            return step2_file_path
            
        except Exception as e:
            logging.error(f"Error processing sheet '{sheet_name}' to Step 2: {str(e)}")
            return None
    
    def process_multiple_sheets_to_step2(self, uploaded_file: UploadedFile, step1_files: list) -> Dict[str, Any]:
        """
        Process multiple Step 1 files to Step 2
        
        Args:
            uploaded_file: Original uploaded Excel file
            step1_files: List of dictionaries with Step 1 file information
            
        Returns:
            Dictionary with processing results
        """
        results = {
            'success_count': 0,
            'failed_count': 0,
            'step2_files': [],
            'failed_sheets': [],
            'total_sheets': len(step1_files)
        }
        
        for step1_info in step1_files:
            sheet_name = step1_info['sheet_name']
            step1_file_path = step1_info['file_path']
            
            try:
                step2_file_path = self.process_sheet_to_step2(uploaded_file, sheet_name, step1_file_path)
                
                if step2_file_path:
                    results['success_count'] += 1
                    results['step2_files'].append({
                        'sheet_name': sheet_name,
                        'filename': Path(step2_file_path).name,
                        'file_path': step2_file_path,
                        'step1_file': step1_file_path
                    })
                else:
                    results['failed_count'] += 1
                    results['failed_sheets'].append(sheet_name)
                    
            except Exception as e:
                logging.error(f"Failed to process sheet '{sheet_name}' to Step 2: {str(e)}")
                results['failed_count'] += 1
                results['failed_sheets'].append(sheet_name)
        
        return results