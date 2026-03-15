"""Step 3 — Data transfer from source sheet to Step 2 template."""

import logging
import openpyxl
import pandas as pd
from pathlib import Path
from typing import List, Optional, Tuple

from streamlit.runtime.uploaded_file_manager import UploadedFile

from .config_manager import ConfigManager


class CoreStep3DataTransfer:
    """Core Step 3 data transfer processor using unified configuration"""

    def __init__(self, output_dir: Optional[str] = None, config_dir: Optional[str] = None):
        self.config_dir = config_dir
        self.processing_config = ConfigManager.get_processing_config(config_dir)

        if output_dir:
            self.output_dir = Path(output_dir)
        else:
            output_dir_name = self.processing_config.get("general", {}).get("output_dir", "output")
            self.output_dir = Path(output_dir_name)

        self.output_dir.mkdir(parents=True, exist_ok=True)

        step3_config = self.processing_config["step3_config"]
        self.header_pattern = step3_config["header_pattern"]
        self.column_mapping = step3_config["column_mapping"]
        self.column_combinations = step3_config.get("column_combinations", {})

    def transfer_data(self, step2_path: str, original_source, sheet_name: str) -> str:
        """
        Transfer data from original Excel sheet to Step 2 template.

        Args:
            step2_path: Path to Step 2 file
            original_source: Either file path (str) or UploadedFile object
            sheet_name: Name of the sheet to transfer data from

        Returns:
            Path to created Step 3 file
        """
        try:
            step3_path = self._get_step3_filename(step2_path)

            wb = openpyxl.load_workbook(step2_path)
            ws = wb.active

            if isinstance(original_source, str):
                header_row, data_rows = self._find_data_in_file(original_source, sheet_name)
            else:
                header_row, data_rows = self._find_data_in_uploaded_file(original_source, sheet_name)

            if header_row is not None and data_rows:
                self._transfer_mapped_data(ws, data_rows, header_row)
            else:
                logging.warning(f"No data found to transfer for sheet '{sheet_name}'")

            self._add_checkbox_markings_step3(ws)

            wb.save(step3_path)
            logging.info(f"Created Step 3 file: {step3_path}")
            return step3_path

        except Exception as e:
            logging.error(f"Error transferring data: {str(e)}")
            raise

    def _find_data_in_file(
        self, file_path: str, sheet_name: str
    ) -> Tuple[Optional[int], List[List]]:
        """Find header row and data rows in local file."""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            return self._find_data_in_dataframe(df)
        except Exception as e:
            logging.error(f"Error reading file data: {str(e)}")
            return None, []

    def _find_data_in_uploaded_file(
        self, uploaded_file: UploadedFile, sheet_name: str
    ) -> Tuple[Optional[int], List[List]]:
        """Find header row and data rows in uploaded file."""
        try:
            df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)
            return self._find_data_in_dataframe(df)
        except Exception as e:
            logging.error(f"Error reading uploaded file data: {str(e)}")
            return None, []

    def _find_data_in_dataframe(
        self, df: pd.DataFrame
    ) -> Tuple[Optional[int], List[List]]:
        """Find header row and extract data rows from DataFrame."""
        header_row = None
        data_rows = []

        # Search for header pattern (case-insensitive)
        for idx, row in df.iterrows():
            for _, cell_value in enumerate(row):
                if pd.isna(cell_value):
                    continue
                if self.header_pattern.upper() in str(cell_value).strip().upper():
                    header_row = idx
                    break
            if header_row is not None:
                break

        if header_row is not None:
            data_start_offset = self.processing_config.get("step3_config", {}).get("data_start_offset", 3)
            actual_data_start_row = header_row + data_start_offset

            for idx in range(actual_data_start_row, len(df)):
                row_data = df.iloc[idx].fillna('').tolist()
                if any(str(cell).strip() for cell in row_data):
                    data_rows.append(row_data)

        return header_row, data_rows

    def _transfer_mapped_data(self, worksheet, data_rows: List[List], header_row: int):
        """Transfer data using column mapping configuration."""
        template_config = ConfigManager.get_template_config()
        start_row = template_config["layout"]["data_start_row"]

        if not data_rows:
            return

        for row_idx, row_data in enumerate(data_rows, start_row):
            for source_col, target_col in self.column_mapping.items():
                try:
                    source_idx = openpyxl.utils.column_index_from_string(source_col) - 1  # 0-based
                    target_idx = openpyxl.utils.column_index_from_string(target_col)       # 1-based

                    if source_idx < len(row_data):
                        source_value = row_data[source_idx]
                        if source_value and str(source_value).strip():
                            worksheet.cell(row_idx, target_idx, source_value)
                except Exception as e:
                    logging.error(f"Error mapping column {source_col} to {target_col}: {str(e)}")
                    continue

        self._apply_column_combinations(worksheet, data_rows, start_row)

    def _apply_column_combinations(
        self, worksheet, data_rows: List[List], start_row: int
    ):
        """Apply column combinations like L+M → I with separator."""
        if not self.column_combinations:
            return

        for row_idx, row_data in enumerate(data_rows, start_row):
            for target_col, config in self.column_combinations.items():
                try:
                    source_cols = config["source_columns"]
                    separator = config["separator"]

                    values = []
                    for source_col in source_cols:
                        source_idx = openpyxl.utils.column_index_from_string(source_col) - 1
                        if source_idx < len(row_data):
                            value = row_data[source_idx]
                            if value and str(value).strip():
                                values.append(str(value).strip())

                    if len(values) == len(source_cols):
                        combined_value = separator.join(values)
                        target_idx = ord(target_col) - ord('A') + 1
                        worksheet.cell(row_idx, target_idx, combined_value)

                except Exception as e:
                    logging.error(f"Error applying combination {target_col}: {str(e)}")
                    continue

    def _add_checkbox_markings_step3(self, worksheet):
        """
        Add X marks in data rows for article columns in Step 3.

        Auto-detects article columns by checking which columns (R, S, T, …)
        have content in rows 1-10.
        """
        try:
            template_config = ConfigManager.get_template_config()
            layout_config = template_config["layout"]
            data_start_row = layout_config.get("data_start_row", 11)
            start_column_letter = layout_config["article_info_rows"]["article_name_start_column"]
            start_column_index = openpyxl.utils.column_index_from_string(start_column_letter)

            article_columns = []
            for col_offset in range(10):
                current_column_index = start_column_index + col_offset
                has_article_info = any(
                    worksheet.cell(row_num, current_column_index).value and
                    str(worksheet.cell(row_num, current_column_index).value).strip()
                    for row_num in range(1, 11)
                )
                if has_article_info:
                    article_columns.append(current_column_index)

            if not article_columns:
                logging.info("No article columns detected, skipping checkbox markings")
                return

            logging.info(f"Detected {len(article_columns)} article column(s) for X markings")

            markings_added = 0
            for row_num in range(data_start_row, worksheet.max_row + 1):
                row_has_data = any(
                    worksheet.cell(row_num, col).value and
                    str(worksheet.cell(row_num, col).value).strip()
                    for col in range(1, 15)
                )
                if not row_has_data:
                    continue

                for column_index in article_columns:
                    worksheet.cell(row_num, column_index, "X")
                    markings_added += 1

            logging.info(f"Added {markings_added} X marks to {len(article_columns)} article column(s)")

        except Exception as e:
            logging.error(f"Error adding checkbox markings in Step 3: {str(e)}")

    def _get_step3_filename(self, step2_path: str) -> str:
        """Generate Step 3 filename from Step 2 path."""
        step2_file = Path(step2_path)
        base_name = step2_file.stem.replace(" - Step2", "")
        filename_template = self.processing_config["general"]["file_naming"]["step3"]

        parts = base_name.split(" - ")
        if len(parts) >= 2:
            original_base = " - ".join(parts[:-1])
            sheet_name = parts[-1]
        else:
            original_base = base_name
            sheet_name = "Unknown"

        step3_filename = filename_template.format(base_name=original_base, sheet_name=sheet_name)
        return str(self.output_dir / step3_filename)
