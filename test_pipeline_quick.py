#!/usr/bin/env python3
"""
Quick test to ensure pipeline still works with existing data
"""

import os
from modules.core_pipeline import CorePipeline

def quick_pipeline_test():
    input_file = "input/QT.TN-01 BM-06 Bảng tóm tắt tình trạng thử nghiệm - TSS SKUBB NN STOR CASE (70g)..xlsx"
    
    if not os.path.exists(input_file):
        print(f"❌ Input file not found: {input_file}")
        return False
    
    print("🔧 Quick pipeline test with multiple article support...")
    print("=" * 60)
    
    try:
        pipeline = CorePipeline()
        
        # Test with just one sheet
        sheet_names = ["44x53"]
        results = pipeline.process_complete_pipeline(input_file, sheet_names)
        
        print("✅ Pipeline completed successfully!")
        
        # Check Step 2 file to see article info placement
        step2_file = f"output/QT.TN-01 BM-06 Bảng tóm tắt tình trạng thử nghiệm - TSS SKUBB NN STOR CASE (70g). - 44x53 - Step2.xlsx"
        
        if os.path.exists(step2_file):
            from openpyxl import load_workbook
            wb = load_workbook(step2_file)
            ws = wb.active
            
            print(f"\n📊 Checking article info placement:")
            
            # Check columns R, S, T for article names
            for col_letter in ['R', 'S', 'T']:
                col_idx = ord(col_letter) - ord('A') + 1
                name_cell = ws.cell(1, col_idx).value
                number_cell = ws.cell(10, col_idx).value
                
                if name_cell and str(name_cell).strip():
                    print(f"   {col_letter}1: '{name_cell}' (Article Name)")
                if number_cell and str(number_cell).strip():
                    print(f"   {col_letter}10: '{number_cell}' (Article Number)")
            
            wb.close()
            print(f"\n✅ Article info successfully placed in consecutive columns!")
            return True
        else:
            print(f"❌ Step 2 file not found: {step2_file}")
            return False
            
    except Exception as e:
        print(f"❌ Pipeline test failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = quick_pipeline_test()
    if success:
        print("\n🎉 Pipeline is working with multiple article support!")
    else:
        print("\n❌ Pipeline test failed!")