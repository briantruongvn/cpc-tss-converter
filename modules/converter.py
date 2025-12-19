"""
Template Converter Module
Integrates the existing step1_template_creation logic for web application use.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from pathlib import Path
from typing import Union, Optional, Dict, Any
import tempfile
import os
from streamlit.runtime.uploaded_file_manager import UploadedFile

class TemplateConverter:
    """
    Template converter that creates Step 1 templates from Excel sheets
    Adapted from the existing step1_template_creation.py for web use
    """
    
    def __init__(self, output_dir: Optional[str] = None):
        """
        Initialize the template converter
        
        Args:
            output_dir: Directory to save output files (default: ./output)
        """
        # Set up output directory
        if output_dir:
            self.output_dir = Path(output_dir)
        else:
            self.output_dir = Path("output")
        
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Template structure configuration
        self.template_headers = self._get_template_headers()
        
        # Define styles (adapted from original code)
        self.row1_2_style = {
            "font": Font(bold=True, color="00000000"),
            "fill": PatternFill(start_color="00B8E6B8", end_color="00B8E6B8", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True)
        }
        
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    def _get_template_headers(self) -> list:
        """
        Get the template headers configuration
        Based on the TSS reference template with exact color coding
        """
        return [
            {"name": "Combination", "font_color": "00000000", "bg_color": "00FFFF00", "width": 15},
            {"name": "General Type Component(Type)", "font_color": "00FFFFFF", "bg_color": "00FF0000", "width": 25},
            {"name": "Sub-Type Component Identity Process Name", "font_color": "00FFFFFF", "bg_color": "00FF0000", "width": 35},
            {"name": "Material Designation", "font_color": "00FFFFFF", "bg_color": "00FF0000", "width": 20},
            {"name": "Material Distributor", "font_color": "00FFFFFF", "bg_color": "00FF0000", "width": 20},
            {"name": "Producer", "font_color": "00FFFFFF", "bg_color": "00FF0000", "width": 15},
            {"name": "Material Type In Process", "font_color": "00FFFFFF", "bg_color": "00FF0000", "width": 22},
            {"name": "Document type", "font_color": "00FFFFFF", "bg_color": "000000FF", "width": 15},
            {"name": "Requirement Source/TED", "font_color": "00FFFFFF", "bg_color": "000000FF", "width": 22},
            {"name": "Sub-type", "font_color": "00FFFFFF", "bg_color": "000000FF", "width": 12},
            {"name": "Regulation or substances", "font_color": "00FFFFFF", "bg_color": "000000FF", "width": 22},
            {"name": "Limit", "font_color": "00000000", "bg_color": "00B8E6B8", "width": 12},
            {"name": "Test method", "font_color": "00000000", "bg_color": "00B8E6B8", "width": 15},
            {"name": "Frequency", "font_color": "00000000", "bg_color": "00B8E6B8", "width": 12},
            {"name": "Level", "font_color": "00FFFFFF", "bg_color": "000000FF", "width": 10},
            {"name": "Warning Limit", "font_color": "00000000", "bg_color": "00B8E6B8", "width": 15},
            {"name": "Additional Information", "font_color": "00000000", "bg_color": "00B8E6B8", "width": 25}
        ]
    
    def create_template_for_sheet(self, sheet_name: str, original_filename: str) -> Optional[str]:
        """
        Create a Step 1 template for a specific sheet
        
        Args:
            sheet_name: Name of the sheet to create template for
            original_filename: Original filename for naming convention
            
        Returns:
            Path to created template file or None if failed
        """
        try:
            # Generate output filename with new format: [Input filename] - [Sheetname] - Step1.xlsx
            base_name = Path(original_filename).stem
            output_filename = f"{base_name} - {sheet_name} - Step1.xlsx"
            output_path = self.output_dir / output_filename
            
            # Create new workbook with template structure
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Template"
            
            # Row 1: Article name with formatting
            cell1 = ws.cell(1, 1, "Article name")
            cell1.font = self.row1_2_style["font"]
            cell1.fill = self.row1_2_style["fill"]
            cell1.alignment = self.row1_2_style["alignment"]
            
            # Row 2: Article number with formatting
            cell2 = ws.cell(2, 1, "Article number")
            cell2.font = self.row1_2_style["font"]
            cell2.fill = self.row1_2_style["fill"]
            cell2.alignment = self.row1_2_style["alignment"]
            
            # Row 3: Headers (17 columns A-Q) with specific formatting and column widths
            for col_idx, header_info in enumerate(self.template_headers, 1):
                cell = ws.cell(3, col_idx, header_info["name"])
                
                # Apply font with specific color for each column
                cell.font = Font(bold=True, color=header_info["font_color"])
                
                # Apply background color
                cell.fill = PatternFill(
                    start_color=header_info["bg_color"], 
                    end_color=header_info["bg_color"], 
                    fill_type="solid"
                )
                
                # Apply alignment
                cell.alignment = self.header_alignment
                
                # Set column width
                col_letter = chr(64 + col_idx)
                ws.column_dimensions[col_letter].width = header_info["width"]
            
            # Save template
            wb.save(str(output_path))
            logging.info(f"Created template: {output_path}")
            
            return str(output_path)
            
        except Exception as e:
            logging.error(f"Error creating template for sheet '{sheet_name}': {str(e)}")
            return None
    
    def create_multiple_templates(self, sheet_names: list, original_filename: str) -> Dict[str, Any]:
        """
        Create templates for multiple sheets
        
        Args:
            sheet_names: List of sheet names to process
            original_filename: Original filename for naming convention
            
        Returns:
            Dictionary with results and file paths
        """
        results = {
            'success_count': 0,
            'failed_count': 0,
            'created_files': [],
            'failed_sheets': [],
            'total_sheets': len(sheet_names)
        }
        
        for sheet_name in sheet_names:
            template_path = self.create_template_for_sheet(sheet_name, original_filename)
            
            if template_path:
                results['success_count'] += 1
                results['created_files'].append({
                    'sheet_name': sheet_name,
                    'filename': f"{sheet_name} - Step1.xlsx",
                    'file_path': template_path
                })
            else:
                results['failed_count'] += 1
                results['failed_sheets'].append(sheet_name)
        
        return results
    
    def validate_template_structure(self, template_path: str) -> bool:
        """
        Validate that a created template has the correct structure
        
        Args:
            template_path: Path to template file to validate
            
        Returns:
            True if valid, False otherwise
        """
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Check row 1
            if ws.cell(1, 1).value != "Article name":
                return False
            
            # Check row 2
            if ws.cell(2, 1).value != "Article number":
                return False
            
            # Check headers (row 3)
            for col_idx, header_info in enumerate(self.template_headers, 1):
                cell_value = ws.cell(3, col_idx).value
                if cell_value != header_info["name"]:
                    return False
            
            return True
            
        except Exception as e:
            logging.error(f"Error validating template structure: {str(e)}")
            return False
    
    def get_output_directory(self) -> str:
        """Get the output directory path"""
        return str(self.output_dir)
    
    def cleanup_old_files(self, max_age_hours: int = 24):
        """
        Clean up old template files from output directory
        
        Args:
            max_age_hours: Maximum age in hours for files to keep
        """
        try:
            import time
            current_time = time.time()
            max_age_seconds = max_age_hours * 3600
            
            for file_path in self.output_dir.glob("*.xlsx"):
                file_age = current_time - file_path.stat().st_mtime
                if file_age > max_age_seconds:
                    file_path.unlink()
                    logging.info(f"Cleaned up old file: {file_path}")
                    
        except Exception as e:
            logging.warning(f"Error during cleanup: {str(e)}")
    
    def get_template_info(self) -> Dict[str, Any]:
        """
        Get information about the template structure
        
        Returns:
            Dictionary with template information
        """
        return {
            'header_count': len(self.template_headers),
            'columns': [h["name"] for h in self.template_headers],
            'structure': {
                'row_1': "Article name",
                'row_2': "Article number", 
                'row_3': "Headers (A-Q)"
            },
            'output_format': "[SheetName] - Step1.xlsx"
        }