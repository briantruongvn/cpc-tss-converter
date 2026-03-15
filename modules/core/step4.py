"""Step 4 — Duplicate removal and data transformations."""

import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from .config_manager import ConfigManager


class CoreStep4DuplicateRemover:
    """Core Step 4 duplicate remover using unified configuration"""

    def __init__(self, output_dir: Optional[str] = None, config_dir: Optional[str] = None):
        self.config_dir = config_dir
        self.processing_config = ConfigManager.get_processing_config(config_dir)

        if output_dir:
            self.output_dir = Path(output_dir)
        else:
            output_dir_name = self.processing_config.get("general", {}).get("output_dir", "output")
            self.output_dir = Path(output_dir_name)

        self.output_dir.mkdir(parents=True, exist_ok=True)

        step4_config = self.processing_config["step4_config"]
        self.comparison_columns = step4_config["comparison_columns"]
        self.max_columns = step4_config["max_columns"]
        self.template_config = ConfigManager.get_template_config(config_dir)

    def remove_duplicates(self, step3_path: str) -> str:
        """
        Remove duplicate rows from Step 3 file based on configured columns.

        Returns path to created Step 4 file.
        """
        try:
            step4_path = self._get_step4_filename(step3_path)

            wb = openpyxl.load_workbook(step3_path)
            ws = wb.active

            # Apply transformations before deduplication
            layout_config = self.template_config["layout"]
            self._apply_data_transformations(ws, layout_config)

            unique_rows = self._extract_unique_rows(ws)
            self._create_deduplicated_file(ws, unique_rows, step4_path)

            logging.info(f"Created Step 4 file: {step4_path}")
            return step4_path

        except Exception as e:
            logging.error(f"Error removing duplicates: {str(e)}")
            raise

    def _extract_unique_rows(self, worksheet) -> List[Tuple[int, List]]:
        """Extract unique rows based on comparison columns, preserving cell objects for formatting."""
        seen_combinations = set()
        unique_rows = []

        template_config = ConfigManager.get_template_config()
        data_start_row = template_config["layout"]["data_start_row"]
        max_col = min(worksheet.max_column, self.max_columns)

        for row_num in range(data_start_row, worksheet.max_row + 1):
            comparison_data = []
            full_row_cells = []

            for col in range(1, max_col + 1):
                cell = worksheet.cell(row_num, col)
                cell_str = str(cell.value).strip() if cell.value is not None else ""

                if col in self.comparison_columns:
                    comparison_data.append(cell_str)

                full_row_cells.append(cell)

            comparison_key = tuple(comparison_data)
            if comparison_key not in seen_combinations:
                seen_combinations.add(comparison_key)
                unique_rows.append((row_num, full_row_cells))

        return unique_rows

    def _create_deduplicated_file(
        self,
        source_worksheet,
        unique_rows: List[Tuple[int, List]],
        output_path: str,
    ):
        """Create new file with deduplicated data."""
        try:
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = self.processing_config["general"]["worksheet_titles"]["step4"]

            template_config = ConfigManager.get_template_config()
            layout_config = template_config["layout"]
            header_row = layout_config["header_row"]
            article_column = openpyxl.utils.column_index_from_string(
                layout_config["article_info_rows"]["article_name_start_column"]
            )

            # Copy all merged cell ranges first
            merge_ranges_copied = 0
            for merge_range in source_worksheet.merged_cells:
                try:
                    new_ws.merge_cells(str(merge_range))
                    merge_ranges_copied += 1
                except Exception as e:
                    logging.warning(f"Could not copy merge range {merge_range}: {e}")

            logging.info(f"Copied {merge_ranges_copied} merged cell ranges to Step 4")

            # Copy header rows with formatting
            for row in range(1, header_row + 1):
                for col in range(1, source_worksheet.max_column + 1):
                    source_cell = source_worksheet.cell(row, col)
                    new_cell = new_ws.cell(row, col)

                    try:
                        new_cell.value = source_cell.value
                    except AttributeError:
                        pass  # MergedCell — skip

                    if source_cell.font:
                        new_cell.font = Font(
                            bold=source_cell.font.bold,
                            color=source_cell.font.color,
                            italic=source_cell.font.italic,
                            underline=source_cell.font.underline
                        )
                    if source_cell.fill:
                        new_cell.fill = PatternFill(
                            start_color=source_cell.fill.start_color,
                            end_color=source_cell.fill.end_color,
                            fill_type=source_cell.fill.fill_type
                        )
                    if source_cell.alignment:
                        if col == article_column and source_cell.alignment.text_rotation:
                            logging.info(
                                f"Copying rotation {source_cell.alignment.text_rotation}° "
                                f"to cell R{row}"
                            )
                        new_cell.alignment = Alignment(
                            horizontal=source_cell.alignment.horizontal,
                            vertical=source_cell.alignment.vertical,
                            wrap_text=source_cell.alignment.wrap_text,
                            text_rotation=source_cell.alignment.text_rotation
                        )
                    if source_cell.border:
                        new_cell.border = Border(
                            left=source_cell.border.left,
                            right=source_cell.border.right,
                            top=source_cell.border.top,
                            bottom=source_cell.border.bottom
                        )

                    if row == 1:
                        col_letter = chr(64 + col)
                        if col_letter in source_worksheet.column_dimensions:
                            new_ws.column_dimensions[col_letter].width = (
                                source_worksheet.column_dimensions[col_letter].width
                            )

            data_start_row = layout_config["data_start_row"]

            # Write unique data rows with formatting
            for new_row_idx, (_, row_cells) in enumerate(unique_rows, data_start_row):
                for col_idx, source_cell in enumerate(row_cells, 1):
                    if source_cell.value:
                        new_cell = new_ws.cell(new_row_idx, col_idx, source_cell.value)

                        if source_cell.font:
                            new_cell.font = Font(
                                bold=source_cell.font.bold,
                                color=source_cell.font.color,
                                italic=source_cell.font.italic,
                                underline=source_cell.font.underline
                            )
                        if source_cell.fill:
                            new_cell.fill = PatternFill(
                                start_color=source_cell.fill.start_color,
                                end_color=source_cell.fill.end_color,
                                fill_type=source_cell.fill.fill_type
                            )
                        if source_cell.alignment:
                            new_cell.alignment = Alignment(
                                horizontal=source_cell.alignment.horizontal,
                                vertical=source_cell.alignment.vertical,
                                wrap_text=source_cell.alignment.wrap_text,
                                text_rotation=source_cell.alignment.text_rotation
                            )
                        if source_cell.border:
                            new_cell.border = Border(
                                left=source_cell.border.left,
                                right=source_cell.border.right,
                                top=source_cell.border.top,
                                bottom=source_cell.border.bottom
                            )

            new_wb.save(output_path)

        except Exception as e:
            logging.error(f"Error creating deduplicated file: {str(e)}")
            raise

    def _apply_data_transformations(self, worksheet, layout_config: dict):
        """Apply data transformations to the worksheet (Step 4 sub-step)."""
        try:
            logging.info("Starting data transformations (Step 4 sub-step)")

            transformations = self.processing_config.get("step4_config", {}).get("transformations", {})
            if not transformations:
                logging.info("No transformations configured, skipping")
                return

            data_start_row = layout_config["data_start_row"]
            column_a_replacements = transformations.get("column_a_replacements", {})
            column_h_k_logic = transformations.get("column_h_k_logic", {})
            transformation_count = 0

            for row_num in range(data_start_row, worksheet.max_row + 1):
                row_has_data = any(
                    worksheet.cell(row_num, col).value and
                    str(worksheet.cell(row_num, col).value).strip()
                    for col in range(1, min(15, worksheet.max_column + 1))
                )
                if not row_has_data:
                    continue

                # Column A text replacements
                col_a_cell = worksheet.cell(row_num, 1)
                if col_a_cell.value and str(col_a_cell.value).strip():
                    original_value = str(col_a_cell.value).strip()
                    lower_value = original_value.lower()
                    for search_term, replacement in column_a_replacements.items():
                        if search_term.lower() == lower_value:
                            logging.info(f"Row {row_num} Col A: '{original_value}' → '{replacement}'")
                            col_a_cell.value = replacement
                            transformation_count += 1
                            break

                # Column H = "SD" → empty Column K
                if column_h_k_logic.get("sd_empty_k", False):
                    col_h_cell = worksheet.cell(row_num, 8)
                    col_k_cell = worksheet.cell(row_num, 11)
                    if (
                        col_h_cell.value and
                        str(col_h_cell.value).strip().upper() == "SD" and
                        col_k_cell.value and
                        str(col_k_cell.value).strip()
                    ):
                        logging.info(
                            f"Row {row_num} Col H=SD: Emptying Col K "
                            f"(was: '{col_k_cell.value}')"
                        )
                        col_k_cell.value = ""
                        transformation_count += 1

            logging.info(f"Data transformations completed: {transformation_count} changes applied")

        except Exception as e:
            logging.error(f"Error applying data transformations: {str(e)}")
            raise

    def _get_step4_filename(self, step3_path: str) -> str:
        """Generate Step 4 filename from Step 3 path."""
        step3_file = Path(step3_path)
        base_name = step3_file.stem.replace(" - Step3", "")
        filename_template = self.processing_config["general"]["file_naming"]["step4"]

        parts = base_name.split(" - ")
        if len(parts) >= 2:
            original_base = " - ".join(parts[:-1])
            sheet_name = parts[-1]
        else:
            original_base = base_name
            sheet_name = "Unknown"

        step4_filename = filename_template.format(base_name=original_base, sheet_name=sheet_name)
        return str(self.output_dir / step4_filename)
