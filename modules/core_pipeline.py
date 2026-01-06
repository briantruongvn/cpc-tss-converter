"""
Core Pipeline Module - Single Point of Truth
Unified business logic for both local and Streamlit pipelines
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border
import pandas as pd
from pathlib import Path
from typing import Union, Optional, Dict, Any, List, Tuple
import logging
import re
import tempfile
from streamlit.runtime.uploaded_file_manager import UploadedFile

class ConfigManager:
    """Manages configuration loading and caching"""
    
    _template_config = None
    _processing_config = None
    
    @classmethod
    def get_template_config(cls, config_dir: Optional[str] = None) -> Dict[str, Any]:
        """Load template configuration from JSON file"""
        if cls._template_config is None:
            if config_dir:
                config_path = Path(config_dir) / "template_config.json"
            else:
                config_path = Path(__file__).parent.parent / "config" / "template_config.json"
            
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    cls._template_config = json.load(f)
            except Exception as e:
                logging.error(f"Failed to load template config: {e}")
                raise
        
        return cls._template_config
    
    @classmethod
    def get_processing_config(cls, config_dir: Optional[str] = None) -> Dict[str, Any]:
        """Load processing configuration from JSON file"""
        if cls._processing_config is None:
            if config_dir:
                config_path = Path(config_dir) / "processing_config.json"
            else:
                config_path = Path(__file__).parent.parent / "config" / "processing_config.json"
            
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    cls._processing_config = json.load(f)
            except Exception as e:
                logging.error(f"Failed to load processing config: {e}")
                raise
        
        return cls._processing_config
    
    @classmethod
    def reload_configs(cls):
        """Force reload of configurations"""
        cls._template_config = None
        cls._processing_config = None


class CoreTemplateCreator:
    """Core template creator using unified configuration"""
    
    def __init__(self, output_dir: Optional[str] = None, config_dir: Optional[str] = None):
        """
        Initialize the core template creator
        
        Args:
            output_dir: Directory to save output files
            config_dir: Directory containing config files
        """
        self.config_dir = config_dir
        self.template_config = ConfigManager.get_template_config(config_dir)
        self.processing_config = ConfigManager.get_processing_config(config_dir)
        
        # Set up output directory
        if output_dir:
            self.output_dir = Path(output_dir)
        else:
            output_dir_name = self.processing_config.get("general", {}).get("output_dir", "output")
            self.output_dir = Path(output_dir_name)
        
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def _create_font_style(self, font_config: Dict[str, Any]) -> Font:
        """Create openpyxl Font object from configuration"""
        return Font(
            bold=font_config.get("bold", False),
            color=font_config.get("color", "00000000")
        )
    
    def _create_fill_style(self, fill_config: Dict[str, Any]) -> PatternFill:
        """Create openpyxl PatternFill object from configuration"""
        return PatternFill(
            start_color=fill_config.get("start_color", "00FFFFFF"),
            end_color=fill_config.get("end_color", "00FFFFFF"),
            fill_type=fill_config.get("fill_type", "solid")
        )
    
    def _create_alignment_style(self, alignment_config: Dict[str, Any]) -> Alignment:
        """Create openpyxl Alignment object from configuration"""
        return Alignment(
            horizontal=alignment_config.get("horizontal", "left"),
            vertical=alignment_config.get("vertical", "center"),
            wrap_text=alignment_config.get("wrap_text", False)
        )
    
    def create_template(self, sheet_name: str, original_filename: str) -> Optional[str]:
        """
        Create a Step 1 template for a specific sheet
        
        Args:
            sheet_name: Name of the sheet to create template for
            original_filename: Original filename for naming convention
            
        Returns:
            Path to created template file or None if failed
        """
        try:
            # Generate output filename
            base_name = Path(original_filename).stem
            filename_template = self.processing_config["general"]["file_naming"]["step1"]
            output_filename = filename_template.format(base_name=base_name, sheet_name=sheet_name)
            output_path = self.output_dir / output_filename
            
            # Create new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.processing_config["general"]["worksheet_titles"]["step1"]
            
            # Apply template structure from config
            template_structure = self.template_config["template_structure"]
            layout_config = self.template_config["layout"]
            
            # Get row positions from layout config
            header_row = layout_config["header_row"]
            
            # Skip creating article info rows in columns A1/A2 since we now place them in column R during Step 2
            # The old row_1 and row_2 configs are kept for legacy compatibility but not used in template creation
            
            # Headers at configured row
            headers = template_structure["headers"]
            header_alignment = self._create_alignment_style(self.template_config["header_alignment"])
            
            for col_idx, header_info in enumerate(headers, 1):
                cell = ws.cell(header_row, col_idx, header_info["name"])
                
                # Apply font with specific color
                cell.font = Font(bold=True, color=header_info["font_color"])
                
                # Apply background color
                cell.fill = PatternFill(
                    start_color=header_info["bg_color"],
                    end_color=header_info["bg_color"],
                    fill_type="solid"
                )
                
                # Apply alignment
                cell.alignment = header_alignment
                
                # Set column width
                col_letter = chr(64 + col_idx)
                ws.column_dimensions[col_letter].width = header_info["width"]
            
            # Save template
            wb.save(str(output_path))
            logging.info(f"Created template: {output_path}")
            
            return str(output_path)
            
        except Exception as e:
            logging.error(f"Error creating template for sheet '{sheet_name}': {str(e)}")
            return None
    
    def get_template_headers(self) -> List[Dict[str, Any]]:
        """Get template headers from configuration"""
        return self.template_config["template_structure"]["headers"]
    
    def validate_template_structure(self, template_path: str) -> bool:
        """
        Validate that a created template has the correct structure
        
        Args:
            template_path: Path to template file to validate
            
        Returns:
            True if valid, False otherwise
        """
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            template_structure = self.template_config["template_structure"]
            layout_config = self.template_config["layout"]
            
            # Get row positions from layout config
            header_row = layout_config["header_row"]
            
            # Skip article info validation since we're using column R now
            # The template structure still has row_1 and row_2 for legacy compatibility
            # but we don't validate them anymore since article info is placed dynamically
            
            # Check headers at configured row
            headers = template_structure["headers"]
            for col_idx, header_info in enumerate(headers, 1):
                cell_value = ws.cell(header_row, col_idx).value
                if cell_value != header_info["name"]:
                    return False
            
            return True
            
        except Exception as e:
            logging.error(f"Error validating template structure: {str(e)}")
            return False


class CoreStep2Processor:
    """Core Step 2 processor using unified configuration"""
    
    def __init__(self, output_dir: Optional[str] = None, config_dir: Optional[str] = None):
        self.config_dir = config_dir
        self.processing_config = ConfigManager.get_processing_config(config_dir)
        
        if output_dir:
            self.output_dir = Path(output_dir)
        else:
            output_dir_name = self.processing_config.get("general", {}).get("output_dir", "output")
            self.output_dir = Path(output_dir_name)
        
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Get patterns from config
        step2_config = self.processing_config["step2_config"]
        self.product_name_patterns = step2_config["product_name_patterns"]
        self.article_number_patterns = step2_config["article_number_patterns"]
    
    def extract_article_info_from_uploaded_file(self, uploaded_file: UploadedFile, sheet_name: str) -> Tuple[List[str], List[str]]:
        """
        Extract article names and numbers from uploaded Streamlit file (supports multiple values)
        
        Args:
            uploaded_file: Streamlit uploaded file object
            sheet_name: Name of the sheet to search in
            
        Returns:
            Tuple of (article_names_list, article_numbers_list)
        """
        try:
            # Read the specific sheet
            df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)
            
            # Search for article information in the dataframe
            return self._search_article_info_in_dataframe(df)
            
        except Exception as e:
            logging.error(f"Error extracting article info from sheet '{sheet_name}': {str(e)}")
            return [], []
    
    def extract_article_info_from_file(self, file_path: str, sheet_name: str) -> Tuple[List[str], List[str]]:
        """
        Extract article names and numbers from local file (supports multiple values)
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to search in
            
        Returns:
            Tuple of (article_names_list, article_numbers_list)
        """
        try:
            # Read the specific sheet
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            # Search for article information in the dataframe
            return self._search_article_info_in_dataframe(df)
            
        except Exception as e:
            logging.error(f"Error extracting article info from sheet '{sheet_name}': {str(e)}")
            return [], []
    
    def _search_article_info_in_dataframe(self, df: pd.DataFrame) -> Tuple[List[str], List[str]]:
        """Search for article information in a pandas DataFrame and return arrays"""
        article_name = None
        article_number = None
        
        # Convert all cells to strings and search (only in rows 1-12, 0-indexed 0-11)
        for row_idx, row in df.iterrows():
            # Only search in rows 1-12 (0-indexed 0-11)
            if row_idx > 11:
                break
                
            for col_idx, cell_value in enumerate(row):
                if pd.isna(cell_value):
                    continue
                
                cell_str = str(cell_value).strip()
                
                # Search for product name
                if article_name is None:
                    for pattern in self.product_name_patterns:
                        if pattern in cell_str:
                            # Try to extract the value after the pattern in same cell
                            article_name = self._extract_value_after_pattern(cell_str, pattern)
                            
                            # If no value found in same cell, check next cell in same row
                            if not article_name and col_idx + 1 < len(row):
                                next_cell_value = row.iloc[col_idx + 1]
                                if not pd.isna(next_cell_value):
                                    article_name = str(next_cell_value).strip()
                            break
                
                # Search for article number
                if article_number is None:
                    for pattern in self.article_number_patterns:
                        if pattern in cell_str:
                            # Try to extract the value after the pattern in same cell
                            article_number = self._extract_value_after_pattern(cell_str, pattern)
                            
                            # If no value found in same cell, check next cell in same row
                            if not article_number and col_idx + 1 < len(row):
                                next_cell_value = row.iloc[col_idx + 1]
                                if not pd.isna(next_cell_value):
                                    article_number = str(next_cell_value).strip()
                            break
                
                # If both found, break early
                if article_name and article_number:
                    break
            
            if article_name and article_number:
                break
        
        # Split multiple values by newlines and semicolons
        import re
        
        article_names = []
        if article_name:
            # Split by newlines and semicolons
            names = re.split(r'[\n;]+', article_name.strip())
            article_names = [name.strip() for name in names if name.strip()]
        
        article_numbers = []
        if article_number:
            # Split by commas, newlines and semicolons
            numbers = re.split(r'[,\n;]+', article_number.strip())
            article_numbers = [num.strip() for num in numbers if num.strip()]
        
        return article_names, article_numbers
    
    def _extract_value_after_pattern(self, text: str, pattern: str) -> Optional[str]:
        """Extract value that comes after a pattern in text"""
        try:
            # Find the pattern in the text
            pattern_index = text.find(pattern)
            if pattern_index == -1:
                return None
            
            # Get text after the pattern
            after_pattern = text[pattern_index + len(pattern):].strip()
            
            # Remove common separators at the start
            while after_pattern and after_pattern[0] in ['::', ':', ' ']:
                after_pattern = after_pattern[1:].strip()
            
            return after_pattern if after_pattern else None
            
        except Exception:
            return None
    
    def populate_template_with_article_info(self, template_path: str, article_names: List[str], article_numbers: List[str]) -> str:
        """
        Populate Step 1 template with extracted article information (supports multiple values)
        
        Args:
            template_path: Path to the Step 1 template file
            article_names: List of article names to populate in consecutive columns
            article_numbers: List of article numbers to populate in consecutive columns
            
        Returns:
            Path to the created Step 2 file
        """
        try:
            # Generate Step 2 filename
            step2_path = self._get_step2_filename(template_path)
            
            # Load template
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Get configuration
            template_config = ConfigManager.get_template_config()
            layout_config = template_config["layout"]["article_info_rows"]
            article_style = template_config["article_info_style"]
            
            # Populate multiple article names in consecutive columns (R, S, T, ...)
            if article_names:
                article_name_row = layout_config["article_name_row"]
                article_name_merge_end = layout_config["article_name_merge_end"]
                start_column_letter = layout_config["article_name_start_column"]
                start_column_index = openpyxl.utils.column_index_from_string(start_column_letter)
                
                for i, article_name in enumerate(article_names):
                    # Calculate current column (R=18, S=19, T=20, ...)
                    current_column_index = start_column_index + i
                    current_column_letter = openpyxl.utils.get_column_letter(current_column_index)
                    
                    # Merge cells for current article name (e.g., R1:R9, S1:S9, T1:T9)
                    merge_range = f"{current_column_letter}{article_name_row}:{current_column_letter}{article_name_merge_end}"
                    ws.merge_cells(merge_range)
                    
                    # Set article name in merged cell
                    cell = ws.cell(article_name_row, current_column_index, article_name)
                    
                    # Debug logging
                    logging.info(f"🔄 Placing article name '{article_name}' in {current_column_letter}{article_name_row} with rotation {article_style['article_name_alignment']['text_rotation']}°")
                    
                    # Apply article name styling (with rotation)
                    cell.fill = PatternFill(
                        start_color=article_style["fill"]["start_color"][2:],  # Remove '00' prefix
                        end_color=article_style["fill"]["end_color"][2:],
                        fill_type=article_style["fill"]["fill_type"]
                    )
                    cell.font = Font(
                        bold=article_style["font"]["bold"],
                        color=article_style["font"]["color"][2:]  # Remove '00' prefix
                    )
                    cell.alignment = Alignment(
                        text_rotation=article_style["article_name_alignment"]["text_rotation"],
                        horizontal=article_style["article_name_alignment"]["horizontal"],
                        vertical=article_style["article_name_alignment"]["vertical"]
                    )
            
            # Populate multiple article numbers in consecutive columns (R10, S10, T10, ...)
            if article_numbers:
                article_number_start_row = layout_config["article_number_start_row"]
                start_column_letter = layout_config["article_number_start_column"]
                start_column_index = openpyxl.utils.column_index_from_string(start_column_letter)
                
                for i, article_number in enumerate(article_numbers):
                    # Calculate current column (R=18, S=19, T=20, ...)
                    current_column_index = start_column_index + i
                    current_column_letter = openpyxl.utils.get_column_letter(current_column_index)
                    
                    # Set article number in current column at start row
                    cell = ws.cell(article_number_start_row, current_column_index, article_number)
                    
                    # Debug logging
                    logging.info(f"🔄 Placing article number '{article_number}' in {current_column_letter}{article_number_start_row}")
                    
                    # Apply article number styling (no rotation)
                    cell.fill = PatternFill(
                        start_color=article_style["fill"]["start_color"][2:],
                        end_color=article_style["fill"]["end_color"][2:],
                        fill_type=article_style["fill"]["fill_type"]
                    )
                    cell.font = Font(
                        bold=article_style["font"]["bold"],
                        color=article_style["font"]["color"][2:]
                    )
                    cell.alignment = Alignment(
                        horizontal=article_style["alignment"]["horizontal"],
                        vertical=article_style["alignment"]["vertical"]
                    )
            
            # Save as Step 2 file
            wb.save(step2_path)
            logging.info(f"Created Step 2 file: {step2_path}")
            
            return step2_path
        
        except Exception as e:
            logging.error(f"Error populating template: {str(e)}")
            raise
    
    def _add_checkbox_markings(self, worksheet, article_names: List[str], article_numbers: List[str], layout_config: dict):
        """
        Add X marks in data rows for columns that contain article names/numbers
        
        Args:
            worksheet: The worksheet to add markings to
            article_names: List of article names (determines which columns need X marks)
            article_numbers: List of article numbers (determines which columns need X marks)
            layout_config: Template layout configuration
        """
        try:
            # Get data start row and article column info
            data_start_row = layout_config.get("data_start_row", 11)  # Row 11 (default)
            start_column_letter = layout_config["article_info_rows"]["article_name_start_column"]  # "R"
            start_column_index = openpyxl.utils.column_index_from_string(start_column_letter)  # R = 18
            
            # Determine how many article columns we have (based on max article names or numbers)
            num_article_columns = max(len(article_names), len(article_numbers))
            
            if num_article_columns == 0:
                logging.info("⚠️ No article info found, skipping checkbox markings")
                return
            
            logging.info(f"✅ Adding X marks to {num_article_columns} article column(s) starting from row {data_start_row}")
            
            # Add X marks to each data row for each article column
            markings_added = 0
            
            for row_num in range(data_start_row, worksheet.max_row + 1):
                # Check if this row has any data (not completely empty)
                row_has_data = False
                for col in range(1, 15):  # Check columns A-N for any data
                    cell = worksheet.cell(row_num, col)
                    if cell.value and str(cell.value).strip():
                        row_has_data = True
                        break
                
                if not row_has_data:
                    continue  # Skip empty rows
                
                # Add X marks to article columns for this data row
                for i in range(num_article_columns):
                    current_column_index = start_column_index + i
                    current_column_letter = openpyxl.utils.get_column_letter(current_column_index)
                    
                    # Set X mark in the cell
                    cell = worksheet.cell(row_num, current_column_index, "X")
                    markings_added += 1
            
            logging.info(f"✅ Added {markings_added} X marks to article columns")
            
        except Exception as e:
            logging.error(f"Error adding checkbox markings: {str(e)}")
            # Don't raise exception, just log and continue
    
    def _get_step2_filename(self, step1_path: str) -> str:
        """Generate Step 2 filename from Step 1 path"""
        step1_file = Path(step1_path)
        base_name = step1_file.stem.replace(" - Step1", "")
        filename_template = self.processing_config["general"]["file_naming"]["step2"]
        
        # Extract original base name and sheet name from Step 1 filename
        parts = base_name.split(" - ")
        if len(parts) >= 2:
            original_base = " - ".join(parts[:-1])
            sheet_name = parts[-1]
        else:
            original_base = base_name
            sheet_name = "Unknown"
        
        step2_filename = filename_template.format(base_name=original_base, sheet_name=sheet_name)
        return str(self.output_dir / step2_filename)


class CoreStep3DataTransfer:
    """Core Step 3 data transfer processor using unified configuration"""
    
    def __init__(self, output_dir: Optional[str] = None, config_dir: Optional[str] = None):
        self.config_dir = config_dir
        self.processing_config = ConfigManager.get_processing_config(config_dir)
        
        if output_dir:
            self.output_dir = Path(output_dir)
        else:
            output_dir_name = self.processing_config.get("general", {}).get("output_dir", "output")
            self.output_dir = Path(output_dir_name)
        
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Get configuration
        step3_config = self.processing_config["step3_config"]
        self.header_pattern = step3_config["header_pattern"]
        self.column_mapping = step3_config["column_mapping"]
        self.column_combinations = step3_config.get("column_combinations", {})
    
    def transfer_data(self, step2_path: str, original_source, sheet_name: str) -> str:
        """
        Transfer data from original Excel sheet to Step 2 template
        
        Args:
            step2_path: Path to Step 2 file
            original_source: Either file path (str) or UploadedFile object
            sheet_name: Name of the sheet to transfer data from
            
        Returns:
            Path to created Step 3 file
        """
        try:
            # Generate Step 3 filename
            step3_path = self._get_step3_filename(step2_path)
            
            # Load Step 2 template
            wb = openpyxl.load_workbook(step2_path)
            ws = wb.active
            
            print(f"🔍 DEBUG: Starting data search for sheet '{sheet_name}'")
            
            # Find header row and data in original source
            if isinstance(original_source, str):
                header_row, data_rows = self._find_data_in_file(original_source, sheet_name)
            else:
                header_row, data_rows = self._find_data_in_uploaded_file(original_source, sheet_name)
            
            print(f"🎯 DEBUG: Search results - Header row: {header_row}, Data rows: {len(data_rows) if data_rows else 0}")
            
            if header_row is not None and data_rows:
                print(f"✅ DEBUG: Found data, proceeding with transfer...")
                # Transfer data using column mapping
                self._transfer_mapped_data(ws, data_rows, header_row)
            else:
                print(f"❌ DEBUG: No data found to transfer!")
                if header_row is None:
                    print(f"   - Header pattern not found")
                if not data_rows:
                    print(f"   - No data rows extracted")
            
            # Add X marks in data rows for article columns
            self._add_checkbox_markings_step3(ws)
            
            # Save as Step 3 file
            wb.save(step3_path)
            logging.info(f"Created Step 3 file: {step3_path}")
            
            return step3_path
            
        except Exception as e:
            logging.error(f"Error transferring data: {str(e)}")
            raise
    
    def _find_data_in_file(self, file_path: str, sheet_name: str) -> Tuple[Optional[int], List[List]]:
        """Find header row and data rows in local file"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            return self._find_data_in_dataframe(df)
        except Exception as e:
            logging.error(f"Error reading file data: {str(e)}")
            return None, []
    
    def _find_data_in_uploaded_file(self, uploaded_file: UploadedFile, sheet_name: str) -> Tuple[Optional[int], List[List]]:
        """Find header row and data rows in uploaded file"""
        try:
            df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)
            return self._find_data_in_dataframe(df)
        except Exception as e:
            logging.error(f"Error reading uploaded file data: {str(e)}")
            return None, []
    
    def _find_data_in_dataframe(self, df: pd.DataFrame) -> Tuple[Optional[int], List[List]]:
        """Find header row and extract data rows from DataFrame"""
        header_row = None
        data_rows = []
        
        print(f"🔍 DEBUG: Searching for pattern '{self.header_pattern}' in dataframe with {len(df)} rows")
        
        # Search for header pattern (case-insensitive)
        for idx, row in df.iterrows():
            for col_idx, cell_value in enumerate(row):
                if pd.isna(cell_value):
                    continue
                cell_str = str(cell_value).strip()
                if self.header_pattern.upper() in cell_str.upper():
                    header_row = idx
                    print(f"✅ DEBUG: Found header pattern at row {idx}, column {col_idx}: '{cell_str}'")
                    break
            if header_row is not None:
                break
        
        if header_row is None:
            print(f"❌ DEBUG: Header pattern '{self.header_pattern}' not found in dataframe")
            # Let's see what's actually in the dataframe
            print("📋 DEBUG: First 10 non-empty cells in dataframe:")
            count = 0
            for idx, row in df.iterrows():
                for col_idx, cell_value in enumerate(row):
                    if pd.notna(cell_value) and str(cell_value).strip():
                        print(f"  Row {idx}, Col {col_idx}: '{str(cell_value).strip()}'")
                        count += 1
                        if count >= 10:
                            break
                if count >= 10:
                    break
        
        # If header found, extract data rows that follow
        if header_row is not None:
            # Get data start offset from configuration
            data_start_offset = self.processing_config.get("step3_config", {}).get("data_start_offset", 3)
            actual_data_start_row = header_row + data_start_offset
            
            print(f"📊 DEBUG: Header found at row {header_row}")
            print(f"🔢 DEBUG: Using data_start_offset: {data_start_offset}")
            print(f"📊 DEBUG: Extracting data rows starting from row {actual_data_start_row} (header + {data_start_offset})")
            
            # Get data rows starting from header + offset
            for idx in range(actual_data_start_row, len(df)):
                row_data = df.iloc[idx].fillna('').tolist()
                # Skip empty rows
                if any(str(cell).strip() for cell in row_data):
                    data_rows.append(row_data)
                    print(f"📝 DEBUG: Added data row {idx}: {[str(cell)[:20] + '...' if len(str(cell)) > 20 else str(cell) for cell in row_data[:5]]}")
            
            print(f"✅ DEBUG: Extracted {len(data_rows)} data rows (skipped {data_start_offset} rows after header)")
        
        return header_row, data_rows
    
    def _transfer_mapped_data(self, worksheet, data_rows: List[List], header_row: int):
        """Transfer data using column mapping configuration"""
        # Get layout configuration to determine data start row
        template_config = ConfigManager.get_template_config()
        layout_config = template_config["layout"]
        start_row = layout_config["data_start_row"]  # Start from configured data row (after headers)
        
        print(f"🔄 DEBUG: Starting data transfer")
        print(f"📊 DEBUG: Data rows to transfer: {len(data_rows)}")
        print(f"📍 DEBUG: Start row configured: {start_row}")
        print(f"🗺️  DEBUG: Column mapping: {self.column_mapping}")
        
        if not data_rows:
            print("❌ DEBUG: No data rows to transfer!")
            return
        
        successful_writes = 0
        failed_writes = 0
        
        for row_idx, row_data in enumerate(data_rows, start_row):
            print(f"📝 DEBUG: Processing row {row_idx} with {len(row_data)} columns")
            row_writes = 0
            
            for source_col, target_col in self.column_mapping.items():
                try:
                    # Convert column letters to indices
                    source_idx = ord(source_col) - ord('A')
                    target_idx = ord(target_col) - ord('A') + 1
                    
                    # Get source data if available
                    if source_idx < len(row_data):
                        source_value = row_data[source_idx]
                        if source_value and str(source_value).strip():
                            print(f"✍️  DEBUG: Writing '{source_value}' to cell ({row_idx}, {target_idx}) [Col {target_col}]")
                            worksheet.cell(row_idx, target_idx, source_value)
                            successful_writes += 1
                            row_writes += 1
                        else:
                            print(f"⚪ DEBUG: Empty source value at source column {source_col} (index {source_idx})")
                    else:
                        print(f"⚠️  DEBUG: Source column {source_col} (index {source_idx}) exceeds row data length {len(row_data)}")
                            
                except Exception as e:
                    print(f"❌ DEBUG: Error mapping column {source_col} to {target_col}: {str(e)}")
                    failed_writes += 1
                    continue
            
            print(f"✅ DEBUG: Row {row_idx} completed: {row_writes} cells written")
        
        print(f"📋 DEBUG: Transfer summary: {successful_writes} successful, {failed_writes} failed")
        
        # Apply column combinations (L+M → I, etc.)
        self._apply_column_combinations(worksheet, data_rows, start_row)
        
        if successful_writes == 0:
            print("🚨 DEBUG: NO DATA WAS WRITTEN! Checking possible issues...")
            print(f"   - Data rows available: {len(data_rows)}")
            print(f"   - First row data sample: {data_rows[0][:10] if data_rows else 'None'}")
            print(f"   - Column mappings: {len(self.column_mapping)}")
            print(f"   - Target start row: {start_row}")
    
    def _apply_column_combinations(self, worksheet, data_rows: List[List], start_row: int):
        """Apply column combinations like L+M → I with separator"""
        if not self.column_combinations:
            print("⚪ DEBUG: No column combinations configured, skipping")
            return
            
        print(f"🔧 DEBUG: Applying column combinations: {list(self.column_combinations.keys())}")
        
        combination_writes = 0
        
        for row_idx, row_data in enumerate(data_rows, start_row):
            for target_col, config in self.column_combinations.items():
                try:
                    source_cols = config["source_columns"]  # e.g., ["L", "M"]
                    separator = config["separator"]  # e.g., "-"
                    
                    # Get values from source columns
                    values = []
                    for source_col in source_cols:
                        source_idx = ord(source_col) - ord('A')
                        if source_idx < len(row_data):
                            value = row_data[source_idx]
                            if value and str(value).strip():
                                values.append(str(value).strip())
                            else:
                                print(f"⚪ DEBUG: Empty value in source column {source_col} for row {row_idx}")
                        else:
                            print(f"⚠️  DEBUG: Source column {source_col} exceeds row data length for row {row_idx}")
                    
                    # Combine values if we have data from all source columns
                    if len(values) == len(source_cols):
                        combined_value = separator.join(values)
                        target_idx = ord(target_col) - ord('A') + 1
                        
                        print(f"🔗 DEBUG: Combining {'+'.join(source_cols)} → '{combined_value}' → Column {target_col}")
                        worksheet.cell(row_idx, target_idx, combined_value)
                        combination_writes += 1
                    else:
                        print(f"⚠️  DEBUG: Incomplete data for combination {'+'.join(source_cols)} in row {row_idx}")
                        
                except Exception as e:
                    print(f"❌ DEBUG: Error applying combination {target_col}: {str(e)}")
                    continue
        
        print(f"📋 DEBUG: Combination summary: {combination_writes} combinations applied")
    
    def _add_checkbox_markings_step3(self, worksheet):
        """
        Add X marks in data rows for article columns in Step 3
        Automatically detects article columns from Step 2 file structure
        """
        try:
            # Get configuration
            template_config = ConfigManager.get_template_config()
            layout_config = template_config["layout"]
            data_start_row = layout_config.get("data_start_row", 11)  # Row 11
            start_column_letter = layout_config["article_info_rows"]["article_name_start_column"]  # "R"
            start_column_index = openpyxl.utils.column_index_from_string(start_column_letter)  # R = 18
            
            # Detect article columns by checking which columns (R, S, T, ...) have content in rows 1-10
            article_columns = []
            
            # Check columns R, S, T, U, V, W, X, Y, Z for article content
            for col_offset in range(10):  # Check up to 10 columns (R through Z+)
                current_column_index = start_column_index + col_offset
                current_column_letter = openpyxl.utils.get_column_letter(current_column_index)
                
                # Check if this column has article info in rows 1-10
                has_article_info = False
                for row_num in range(1, 11):  # Check rows 1-10
                    cell = worksheet.cell(row_num, current_column_index)
                    if cell.value and str(cell.value).strip():
                        has_article_info = True
                        break
                
                if has_article_info:
                    article_columns.append(current_column_index)
            
            if not article_columns:
                logging.info("⚠️ No article columns detected, skipping checkbox markings")
                return
            
            logging.info(f"✅ Detected {len(article_columns)} article column(s) for X markings")
            
            # Add X marks to each data row for each detected article column
            markings_added = 0
            
            for row_num in range(data_start_row, worksheet.max_row + 1):
                # Check if this row has any data (not completely empty)
                row_has_data = False
                for col in range(1, 15):  # Check columns A-N for any data
                    cell = worksheet.cell(row_num, col)
                    if cell.value and str(cell.value).strip():
                        row_has_data = True
                        break
                
                if not row_has_data:
                    continue  # Skip empty rows
                
                # Add X marks to detected article columns for this data row
                for column_index in article_columns:
                    cell = worksheet.cell(row_num, column_index, "X")
                    markings_added += 1
            
            logging.info(f"✅ Added {markings_added} X marks to {len(article_columns)} article column(s)")
            
        except Exception as e:
            logging.error(f"Error adding checkbox markings in Step 3: {str(e)}")
            # Don't raise exception, just log and continue
    
    def _get_step3_filename(self, step2_path: str) -> str:
        """Generate Step 3 filename from Step 2 path"""
        step2_file = Path(step2_path)
        base_name = step2_file.stem.replace(" - Step2", "")
        filename_template = self.processing_config["general"]["file_naming"]["step3"]
        
        # Extract original base name and sheet name
        parts = base_name.split(" - ")
        if len(parts) >= 2:
            original_base = " - ".join(parts[:-1])
            sheet_name = parts[-1]
        else:
            original_base = base_name
            sheet_name = "Unknown"
        
        step3_filename = filename_template.format(base_name=original_base, sheet_name=sheet_name)
        return str(self.output_dir / step3_filename)


class CoreStep4DuplicateRemover:
    """Core Step 4 duplicate remover using unified configuration"""
    
    def __init__(self, output_dir: Optional[str] = None, config_dir: Optional[str] = None):
        self.config_dir = config_dir
        self.processing_config = ConfigManager.get_processing_config(config_dir)
        
        if output_dir:
            self.output_dir = Path(output_dir)
        else:
            output_dir_name = self.processing_config.get("general", {}).get("output_dir", "output")
            self.output_dir = Path(output_dir_name)
        
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Get configuration
        step4_config = self.processing_config["step4_config"]
        self.comparison_columns = step4_config["comparison_columns"]
        self.max_columns = step4_config["max_columns"]
        
        # Load template config for transformations
        self.template_config = ConfigManager.get_template_config(config_dir)
    
    def remove_duplicates(self, step3_path: str) -> str:
        """
        Remove duplicate rows from Step 3 file based on configured columns
        
        Args:
            step3_path: Path to Step 3 file
            
        Returns:
            Path to created Step 4 file
        """
        try:
            # Generate Step 4 filename
            step4_path = self._get_step4_filename(step3_path)
            
            # Load Step 3 file
            wb = openpyxl.load_workbook(step3_path)
            ws = wb.active
            
            # Apply data transformations FIRST (before duplicate removal)
            layout_config = self.template_config["layout"]
            self._apply_data_transformations(ws, layout_config)
            
            # Extract and deduplicate data (after transformations)
            unique_rows = self._extract_unique_rows(ws)
            
            # Create new workbook with unique data
            self._create_deduplicated_file(ws, unique_rows, step4_path)
            
            logging.info(f"Created Step 4 file: {step4_path}")
            return step4_path
            
        except Exception as e:
            logging.error(f"Error removing duplicates: {str(e)}")
            raise
    
    def _extract_unique_rows(self, worksheet) -> List[Tuple[int, List]]:
        """Extract unique rows based on comparison columns, preserving cell objects for formatting"""
        seen_combinations = set()
        unique_rows = []
        
        # Get layout configuration to determine data start row
        template_config = ConfigManager.get_template_config()
        layout_config = template_config["layout"]
        data_start_row = layout_config["data_start_row"]
        
        # Process rows starting from configured data row (after headers)
        for row_num in range(data_start_row, worksheet.max_row + 1):
            # Extract comparison data from configured columns
            comparison_data = []
            full_row_cells = []
            
            # Get data from all columns up to max columns
            max_col = min(worksheet.max_column, self.max_columns)
            
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row_num, col)
                cell_str = str(cell.value).strip() if cell.value is not None else ""
                
                # Add to comparison data if within comparison columns
                if col in self.comparison_columns:
                    comparison_data.append(cell_str)
                
                # Store cell object to preserve formatting
                full_row_cells.append(cell)
            
            # Create comparison key from columns A-K
            comparison_key = tuple(comparison_data)
            
            # Check if this combination is unique
            if comparison_key not in seen_combinations:
                seen_combinations.add(comparison_key)
                unique_rows.append((row_num, full_row_cells))
        
        return unique_rows
    
    def _create_deduplicated_file(self, source_worksheet, unique_rows: List[Tuple[int, List]], output_path: str):
        """Create new file with deduplicated data"""
        try:
            # Create new workbook
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = self.processing_config["general"]["worksheet_titles"]["step4"]
            
            # Get layout configuration to determine header range
            template_config = ConfigManager.get_template_config()  
            layout_config = template_config["layout"]
            header_row = layout_config["header_row"]
            
            # Get article column for debugging
            article_column = openpyxl.utils.column_index_from_string(layout_config["article_info_rows"]["article_name_start_column"])
            
            # Copy header rows (1 to header_row) from source INCLUDING MERGED CELLS
            # First, copy all merged cell ranges from source to destination
            merge_ranges_copied = 0
            for merge_range in source_worksheet.merged_cells:
                try:
                    # Convert merge range to string and apply to new worksheet
                    new_ws.merge_cells(str(merge_range))
                    merge_ranges_copied += 1
                except Exception as e:
                    logging.warning(f"Could not copy merge range {merge_range}: {e}")
            
            logging.info(f"📋 Copied {merge_ranges_copied} merged cell ranges to Step 4")
            
            # Then copy cell values and formatting
            for row in range(1, header_row + 1):
                for col in range(1, source_worksheet.max_column + 1):
                    source_cell = source_worksheet.cell(row, col)
                    new_cell = new_ws.cell(row, col)
                    
                    # Copy value (handle merged cells properly)
                    try:
                        new_cell.value = source_cell.value
                    except AttributeError:
                        # MergedCell doesn't have value attribute, skip it
                        pass
                    
                    # Copy formatting
                    if source_cell.font:
                        new_cell.font = Font(
                            bold=source_cell.font.bold,
                            color=source_cell.font.color,
                            italic=source_cell.font.italic,
                            underline=source_cell.font.underline
                        )
                    if source_cell.fill:
                        new_cell.fill = PatternFill(
                            start_color=source_cell.fill.start_color,
                            end_color=source_cell.fill.end_color,
                            fill_type=source_cell.fill.fill_type
                        )
                    if source_cell.alignment:
                        # Log rotation for column R cells
                        if col == article_column and source_cell.alignment.text_rotation:
                            logging.info(f"🔄 Copying rotation {source_cell.alignment.text_rotation}° to cell R{row}")
                        
                        new_cell.alignment = Alignment(
                            horizontal=source_cell.alignment.horizontal,
                            vertical=source_cell.alignment.vertical,
                            wrap_text=source_cell.alignment.wrap_text,
                            text_rotation=source_cell.alignment.text_rotation
                        )
                    
                    # Copy border if exists
                    if source_cell.border:
                        new_cell.border = Border(
                            left=source_cell.border.left,
                            right=source_cell.border.right,
                            top=source_cell.border.top,
                            bottom=source_cell.border.bottom
                        )
                    
                    # Copy column width
                    if row == 1:
                        col_letter = chr(64 + col)
                        if col_letter in source_worksheet.column_dimensions:
                            new_ws.column_dimensions[col_letter].width = source_worksheet.column_dimensions[col_letter].width
            
            # Get layout configuration for data start row
            data_start_row = layout_config["data_start_row"]
            
            # Write unique data starting from configured data row WITH FORMATTING
            for new_row_idx, (original_row, row_cells) in enumerate(unique_rows, data_start_row):
                for col_idx, source_cell in enumerate(row_cells, 1):
                    if source_cell.value:  # Only write non-empty values
                        # Create new cell and copy value
                        new_cell = new_ws.cell(new_row_idx, col_idx, source_cell.value)
                        
                        # Copy ALL formatting from source cell
                        if source_cell.font:
                            new_cell.font = Font(
                                bold=source_cell.font.bold,
                                color=source_cell.font.color,
                                italic=source_cell.font.italic,
                                underline=source_cell.font.underline
                            )
                        
                        if source_cell.fill:
                            new_cell.fill = PatternFill(
                                start_color=source_cell.fill.start_color,
                                end_color=source_cell.fill.end_color,
                                fill_type=source_cell.fill.fill_type
                            )
                        
                        if source_cell.alignment:
                            new_cell.alignment = Alignment(
                                horizontal=source_cell.alignment.horizontal,
                                vertical=source_cell.alignment.vertical,
                                wrap_text=source_cell.alignment.wrap_text,
                                text_rotation=source_cell.alignment.text_rotation
                            )
                        
                        if source_cell.border:
                            new_cell.border = Border(
                                left=source_cell.border.left,
                                right=source_cell.border.right,
                                top=source_cell.border.top,
                                bottom=source_cell.border.bottom
                            )
            
            # Save new file
            new_wb.save(output_path)
            
        except Exception as e:
            logging.error(f"Error creating deduplicated file: {str(e)}")
            raise
    
    def _apply_data_transformations(self, worksheet, layout_config: dict):
        """Apply data transformations to the worksheet (sub step of Step 4)"""
        try:
            logging.info("🔧 Starting data transformations (Step 4 sub step)")
            
            # Get transformation configuration
            transformations = self.processing_config.get("step4_config", {}).get("transformations", {})
            if not transformations:
                logging.info("⚠️ No transformations configured, skipping")
                return
            
            data_start_row = layout_config["data_start_row"]
            
            # Get transformation rules
            column_a_replacements = transformations.get("column_a_replacements", {})
            column_h_k_logic = transformations.get("column_h_k_logic", {})
            
            transformation_count = 0
            
            # Process all data rows
            for row_num in range(data_start_row, worksheet.max_row + 1):
                row_has_data = False
                
                # Check if this row has any data
                for col in range(1, min(15, worksheet.max_column + 1)):
                    cell = worksheet.cell(row_num, col)
                    if cell.value and str(cell.value).strip():
                        row_has_data = True
                        break
                
                if not row_has_data:
                    continue
                
                # Column A transformations (case insensitive)
                col_a_cell = worksheet.cell(row_num, 1)  # Column A
                if col_a_cell.value and str(col_a_cell.value).strip():
                    original_value = str(col_a_cell.value).strip()
                    lower_value = original_value.lower()
                    
                    for search_term, replacement in column_a_replacements.items():
                        if search_term.lower() == lower_value:
                            logging.info(f"🔄 Row {row_num} Col A: '{original_value}' → '{replacement}'")
                            col_a_cell.value = replacement
                            transformation_count += 1
                            break
                
                # Column H + K logic
                if column_h_k_logic.get("sd_empty_k", False):
                    col_h_cell = worksheet.cell(row_num, 8)  # Column H
                    col_k_cell = worksheet.cell(row_num, 11)  # Column K
                    
                    if (col_h_cell.value and str(col_h_cell.value).strip().upper() == "SD" and
                        col_k_cell.value and str(col_k_cell.value).strip()):
                        logging.info(f"🔄 Row {row_num} Col H=SD: Emptying Col K (was: '{col_k_cell.value}')")
                        col_k_cell.value = ""
                        transformation_count += 1
            
            logging.info(f"✅ Data transformations completed: {transformation_count} changes applied")
            
        except Exception as e:
            logging.error(f"Error applying data transformations: {str(e)}")
            raise
    
    def _get_step4_filename(self, step3_path: str) -> str:
        """Generate Step 4 filename from Step 3 path"""
        step3_file = Path(step3_path)
        base_name = step3_file.stem.replace(" - Step3", "")
        filename_template = self.processing_config["general"]["file_naming"]["step4"]
        
        # Extract original base name and sheet name
        parts = base_name.split(" - ")
        if len(parts) >= 2:
            original_base = " - ".join(parts[:-1])
            sheet_name = parts[-1]
        else:
            original_base = base_name
            sheet_name = "Unknown"
        
        step4_filename = filename_template.format(base_name=original_base, sheet_name=sheet_name)
        return str(self.output_dir / step4_filename)


class CorePipeline:
    """Unified pipeline orchestrator"""
    
    def __init__(self, output_dir: Optional[str] = None, config_dir: Optional[str] = None):
        """
        Initialize the core pipeline
        
        Args:
            output_dir: Directory for output files
            config_dir: Directory containing configuration files
        """
        self.output_dir = output_dir
        self.config_dir = config_dir
        
        # Initialize all processors
        self.template_creator = CoreTemplateCreator(output_dir, config_dir)
        self.step2_processor = CoreStep2Processor(output_dir, config_dir)
        self.step3_processor = CoreStep3DataTransfer(output_dir, config_dir)
        self.step4_processor = CoreStep4DuplicateRemover(output_dir, config_dir)
    
    def process_complete_pipeline(self, source_file, sheet_names: List[str]) -> Dict[str, Any]:
        """
        Run the complete pipeline for multiple sheets
        
        Args:
            source_file: Either file path (str) or UploadedFile object
            sheet_names: List of sheet names to process
            
        Returns:
            Dictionary with processing results
        """
        results = {
            'success_count': 0,
            'failed_count': 0,
            'step4_files': [],
            'failed_sheets': [],
            'total_sheets': len(sheet_names)
        }
        
        # Determine filename for template creation
        if isinstance(source_file, str):
            original_filename = Path(source_file).name
        else:
            original_filename = source_file.name
        
        for sheet_name in sheet_names:
            try:
                # Step 1: Create template
                template_path = self.template_creator.create_template(sheet_name, original_filename)
                if not template_path:
                    raise Exception("Failed to create template")
                
                # Step 2: Extract and populate article info
                article_name, article_number = self.step2_processor.extract_article_info_from_uploaded_file(source_file, sheet_name) if isinstance(source_file, UploadedFile) else self.step2_processor.extract_article_info_from_file(source_file, sheet_name)
                step2_path = self.step2_processor.populate_template_with_article_info(template_path, article_name, article_number)
                
                # Step 3: Transfer data
                step3_path = self.step3_processor.transfer_data(step2_path, source_file, sheet_name)
                
                # Step 4: Remove duplicates
                step4_path = self.step4_processor.remove_duplicates(step3_path)
                
                results['success_count'] += 1
                results['step4_files'].append({
                    'sheet_name': sheet_name,
                    'filename': Path(step4_path).name,
                    'file_path': step4_path
                })
                
            except Exception as e:
                logging.error(f"Pipeline failed for sheet '{sheet_name}': {str(e)}")
                results['failed_count'] += 1
                results['failed_sheets'].append(sheet_name)
        
        return results
    
    def reload_configuration(self):
        """Reload all configurations"""
        ConfigManager.reload_configs()
        
        # Reinitialize all processors
        self.template_creator = CoreTemplateCreator(self.output_dir, self.config_dir)
        self.step2_processor = CoreStep2Processor(self.output_dir, self.config_dir)
        self.step3_processor = CoreStep3DataTransfer(self.output_dir, self.config_dir)
        self.step4_processor = CoreStep4DuplicateRemover(self.output_dir, self.config_dir)