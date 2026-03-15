"""Step 1 — Template creator."""

import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from typing import Any, Dict, List, Optional

from .config_manager import ConfigManager


class CoreTemplateCreator:
    """Core template creator using unified configuration"""

    def __init__(self, output_dir: Optional[str] = None, config_dir: Optional[str] = None):
        self.config_dir = config_dir
        self.template_config = ConfigManager.get_template_config(config_dir)
        self.processing_config = ConfigManager.get_processing_config(config_dir)

        if output_dir:
            self.output_dir = Path(output_dir)
        else:
            output_dir_name = self.processing_config.get("general", {}).get("output_dir", "output")
            self.output_dir = Path(output_dir_name)

        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _create_font_style(self, font_config: Dict[str, Any]) -> Font:
        """Create openpyxl Font object from configuration"""
        return Font(
            bold=font_config.get("bold", False),
            color=font_config.get("color", "00000000")
        )

    def _create_fill_style(self, fill_config: Dict[str, Any]) -> PatternFill:
        """Create openpyxl PatternFill object from configuration"""
        return PatternFill(
            start_color=fill_config.get("start_color", "00FFFFFF"),
            end_color=fill_config.get("end_color", "00FFFFFF"),
            fill_type=fill_config.get("fill_type", "solid")
        )

    def _create_alignment_style(self, alignment_config: Dict[str, Any]) -> Alignment:
        """Create openpyxl Alignment object from configuration"""
        return Alignment(
            horizontal=alignment_config.get("horizontal", "left"),
            vertical=alignment_config.get("vertical", "center"),
            wrap_text=alignment_config.get("wrap_text", False)
        )

    def create_template(self, sheet_name: str, original_filename: str) -> Optional[str]:
        """
        Create a Step 1 template for a specific sheet.

        Returns path to created template file or None if failed.
        """
        try:
            base_name = Path(original_filename).stem
            filename_template = self.processing_config["general"]["file_naming"]["step1"]
            output_filename = filename_template.format(base_name=base_name, sheet_name=sheet_name)
            output_path = self.output_dir / output_filename

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.processing_config["general"]["worksheet_titles"]["step1"]

            template_structure = self.template_config["template_structure"]
            layout_config = self.template_config["layout"]
            header_row = layout_config["header_row"]

            # Article info rows (row_1/row_2) are placed dynamically in Step 2
            headers = template_structure["headers"]
            header_alignment = self._create_alignment_style(self.template_config["header_alignment"])

            for col_idx, header_info in enumerate(headers, 1):
                cell = ws.cell(header_row, col_idx, header_info["name"])
                cell.font = Font(bold=True, color=header_info["font_color"])
                cell.fill = PatternFill(
                    start_color=header_info["bg_color"],
                    end_color=header_info["bg_color"],
                    fill_type="solid"
                )
                cell.alignment = header_alignment
                col_letter = chr(64 + col_idx)
                ws.column_dimensions[col_letter].width = header_info["width"]

            wb.save(str(output_path))
            logging.info(f"Created template: {output_path}")
            return str(output_path)

        except Exception as e:
            logging.error(f"Error creating template for sheet '{sheet_name}': {str(e)}")
            return None

    def get_template_headers(self) -> List[Dict[str, Any]]:
        """Get template headers from configuration"""
        return self.template_config["template_structure"]["headers"]

    def validate_template_structure(self, template_path: str) -> bool:
        """Validate that a created template has the correct structure."""
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            template_structure = self.template_config["template_structure"]
            layout_config = self.template_config["layout"]
            header_row = layout_config["header_row"]

            headers = template_structure["headers"]
            for col_idx, header_info in enumerate(headers, 1):
                cell_value = ws.cell(header_row, col_idx).value
                if cell_value != header_info["name"]:
                    return False

            return True

        except Exception as e:
            logging.error(f"Error validating template structure: {str(e)}")
            return False
