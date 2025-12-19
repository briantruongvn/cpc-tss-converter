"""
Step 3 Data Transfer Module
Transfers data from original Excel sheets to Step 2 templates using specific column mapping rules.
"""

import openpyxl
from pathlib import Path
from typing import Optional, Dict, Any, Tuple, List
import logging
import re
from streamlit.runtime.uploaded_file_manager import UploadedFile
import tempfile

class Step3DataTransfer:
    """
    Processes Step 2 templates by transferring data from original Excel sheets
    using specific column mapping rules
    """
    
    def __init__(self, output_dir: Optional[str] = None):
        """
        Initialize the Step 3 data transfer processor
        
        Args:
            output_dir: Directory to save Step 3 files (default: ./output)
        """
        if output_dir:
            self.output_dir = Path(output_dir)
        else:
            self.output_dir = Path("output")
        
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Header search pattern
        self.header_pattern = "CẤU THÀNH SẢN PHẨM"
        
        # Column mapping: source_column -> target_column
        self.column_mapping = {
            'A': 'A',  # A -> A
            'B': 'B',  # B -> B
            'C': 'C',  # C -> C
            'D': 'D',  # D -> D
            'I': 'E',  # I -> E
            'J': 'F',  # J -> F
            'S': 'H',  # S -> H
            'N': 'J',  # N -> J
            'O': 'K',  # O -> K
            'P': 'L',  # P -> L
            'R': 'M',  # R -> M
            'Z': 'N',  # Z -> N
            'T': 'O',  # T -> O
            'Q': 'P'   # Q -> P
        }
        
        # Special mapping for L-M combination -> I
        self.lm_combination_target = 'I'
    
    def find_header_row(self, worksheet) -> Optional[int]:
        """
        Find the row containing the header "CẤU THÀNH SẢN PHẨM"
        
        Args:
            worksheet: openpyxl worksheet object
            
        Returns:
            Row number (1-based) if found, None otherwise
        """
        try:
            for row in range(1, min(50, worksheet.max_row + 1)):  # Search first 50 rows
                for col in range(1, min(30, worksheet.max_column + 1)):  # Search first 30 columns
                    cell_value = worksheet.cell(row, col).value
                    
                    if cell_value and isinstance(cell_value, str):
                        cell_text = cell_value.strip()
                        
                        # Check if cell contains the header pattern
                        if self.header_pattern in cell_text:
                            logging.info(f"Found header pattern '{self.header_pattern}' at row {row}, col {col}: '{cell_text}'")
                            
                            # Debug: Examine the structure around found header
                            logging.info("=== DEBUG: Examining rows around found header ===")
                            for debug_row in range(max(1, row - 1), min(worksheet.max_row + 1, row + 6)):
                                row_content = []
                                for debug_col in range(1, min(10, worksheet.max_column + 1)):  # Check first 10 columns
                                    debug_cell = worksheet.cell(debug_row, debug_col)
                                    debug_value = debug_cell.value
                                    if debug_value is not None:
                                        row_content.append(f"Col{debug_col}='{str(debug_value).strip()}'")
                                
                                if row_content:
                                    logging.info(f"Row {debug_row}: {', '.join(row_content)}")
                                else:
                                    logging.info(f"Row {debug_row}: (empty)")
                            
                            logging.info("=== END DEBUG ===")
                            return row
            
            logging.warning(f"Header pattern '{self.header_pattern}' not found, defaulting to row 13")
            return 13  # Default fallback
            
        except Exception as e:
            logging.error(f"Error finding header row: {str(e)}")
            return 13  # Default fallback
    
    def get_data_start_row(self, header_row: int) -> int:
        """
        Calculate the data start row (header row + 4)
        
        Args:
            header_row: Row number of the header
            
        Returns:
            Data start row number
        """
        return header_row + 4
    
    def column_letter_to_number(self, column_letter: str) -> int:
        """
        Convert column letter to number (A=1, B=2, etc.)
        
        Args:
            column_letter: Column letter (e.g., 'A', 'B', 'AB')
            
        Returns:
            Column number
        """
        result = 0
        for char in column_letter:
            result = result * 26 + (ord(char.upper()) - ord('A') + 1)
        return result
    
    def column_number_to_letter(self, column_number: int) -> str:
        """
        Convert column number to letter (1=A, 2=B, etc.)
        
        Args:
            column_number: Column number
            
        Returns:
            Column letter
        """
        result = ""
        while column_number > 0:
            column_number -= 1
            result = chr(column_number % 26 + ord('A')) + result
            column_number //= 26
        return result
    
    def extract_data_from_original_sheet(self, uploaded_file: UploadedFile, sheet_name: str) -> Tuple[Optional[int], List[Dict[str, Any]]]:
        """
        Extract data from the original sheet starting from the header + 4 row
        
        Args:
            uploaded_file: Streamlit uploaded file object
            sheet_name: Name of the sheet to process
            
        Returns:
            Tuple of (data_start_row, list_of_row_data_dicts)
        """
        try:
            # Save uploaded file temporarily
            with tempfile.NamedTemporaryFile(delete=False, suffix=Path(uploaded_file.name).suffix) as tmp_file:
                tmp_file.write(uploaded_file.getvalue())
                tmp_file.flush()
                
                # Load workbook and get specific sheet
                wb = openpyxl.load_workbook(tmp_file.name, data_only=True)
                
                if sheet_name not in wb.sheetnames:
                    logging.error(f"Sheet '{sheet_name}' not found in workbook")
                    return None, []
                
                ws = wb[sheet_name]
                
                # Find header row
                header_row = self.find_header_row(ws)
                if header_row is None:
                    logging.error(f"Could not find header row in sheet '{sheet_name}'")
                    return None, []
                
                # Calculate data start row
                data_start_row = self.get_data_start_row(header_row)
                logging.info(f"Header found at row {header_row}, calculated data start row: {data_start_row}")
                
                # Debug: Show what's at the data start row to verify it's not headers
                logging.info("=== DEBUG: Checking data start row content ===")
                start_row_content = []
                for debug_col in range(1, min(10, ws.max_column + 1)):
                    cell_value = ws.cell(data_start_row, debug_col).value
                    if cell_value is not None:
                        start_row_content.append(f"Col{debug_col}='{str(cell_value).strip()}'")
                
                if start_row_content:
                    logging.info(f"Data start row {data_start_row}: {', '.join(start_row_content)}")
                else:
                    logging.info(f"Data start row {data_start_row}: (empty)")
                logging.info("=== END DEBUG ===")
                
                # Extract data rows
                data_rows = []
                max_col = max([self.column_letter_to_number(col) for col in self.column_mapping.keys()] + [26])  # Z=26
                
                for row in range(data_start_row, ws.max_row + 1):
                    row_data = {}
                    has_data = False
                    
                    # Read all relevant columns for this row
                    for col_num in range(1, max_col + 1):
                        col_letter = self.column_number_to_letter(col_num)
                        cell_value = ws.cell(row, col_num).value
                        
                        if cell_value is not None:
                            row_data[col_letter] = str(cell_value).strip()
                            has_data = True
                        else:
                            row_data[col_letter] = ""
                    
                    # Only add row if it has any data
                    if has_data:
                        data_rows.append(row_data)
                    elif len(data_rows) > 0:
                        # Stop if we encounter empty row after having data
                        break
                
                # Clean up temporary file
                Path(tmp_file.name).unlink()
                
                logging.info(f"Extracted {len(data_rows)} data rows from sheet '{sheet_name}'")
                
                return data_start_row, data_rows
                
        except Exception as e:
            logging.error(f"Error extracting data from sheet '{sheet_name}': {str(e)}")
            return None, []
    
    def apply_column_mapping(self, source_row_data: Dict[str, Any]) -> Dict[str, str]:
        """
        Apply column mapping rules to transform source data to target format
        
        Args:
            source_row_data: Dictionary with source column data
            
        Returns:
            Dictionary with mapped target column data
        """
        mapped_data = {}
        
        try:
            # Apply standard column mappings
            for source_col, target_col in self.column_mapping.items():
                source_value = source_row_data.get(source_col, "")
                mapped_data[target_col] = source_value
            
            # Special handling for L-M combination -> I
            l_value = source_row_data.get('L', "").strip()
            m_value = source_row_data.get('M', "").strip()
            
            if l_value or m_value:
                if l_value and m_value:
                    combined_value = f"{l_value}-{m_value}"
                elif l_value:
                    combined_value = l_value
                else:
                    combined_value = m_value
                
                mapped_data[self.lm_combination_target] = combined_value
                logging.debug(f"Combined L-M: '{l_value}' + '{m_value}' = '{combined_value}'")
            else:
                mapped_data[self.lm_combination_target] = ""
            
            return mapped_data
            
        except Exception as e:
            logging.error(f"Error applying column mapping: {str(e)}")
            return {}
    
    def populate_step2_template_with_data(self, step2_file_path: str, mapped_data_rows: List[Dict[str, str]]) -> str:
        """
        Populate Step 2 template with mapped data and save as Step 3
        
        Args:
            step2_file_path: Path to Step 2 template file
            mapped_data_rows: List of dictionaries with mapped column data
            
        Returns:
            Path to created Step 3 file
        """
        try:
            # Load Step 2 template
            wb = openpyxl.load_workbook(step2_file_path)
            ws = wb.active
            
            # Start writing data from row 4 (after headers)
            start_row = 4
            
            for i, row_data in enumerate(mapped_data_rows):
                current_row = start_row + i
                
                # Write each mapped column to the appropriate cell
                for target_col, value in row_data.items():
                    if value:  # Only write non-empty values
                        target_col_num = self.column_letter_to_number(target_col)
                        cell = ws.cell(current_row, target_col_num)
                        cell.value = value
                        
                        logging.debug(f"Row {current_row}, Col {target_col}: '{value}'")
            
            # Generate Step 3 filename
            step2_path = Path(step2_file_path)
            step3_filename = step2_path.name.replace("Step2.xlsx", "Step3.xlsx")
            step3_path = self.output_dir / step3_filename
            
            # Save as Step 3 file
            wb.save(str(step3_path))
            logging.info(f"Created Step 3 file: {step3_path} with {len(mapped_data_rows)} data rows")
            
            return str(step3_path)
            
        except Exception as e:
            logging.error(f"Error populating Step 2 template with data: {str(e)}")
            raise Exception(f"Failed to create Step 3 file: {str(e)}")
    
    def process_sheet_to_step3(self, uploaded_file: UploadedFile, sheet_name: str, step2_file_path: str) -> Optional[str]:
        """
        Complete Step 3 processing for a single sheet
        
        Args:
            uploaded_file: Original uploaded Excel file
            sheet_name: Name of the sheet to process
            step2_file_path: Path to the Step 2 template file
            
        Returns:
            Path to created Step 3 file, or None if failed
        """
        try:
            # Extract data from original sheet
            data_start_row, data_rows = self.extract_data_from_original_sheet(uploaded_file, sheet_name)
            
            if not data_rows:
                logging.warning(f"No data rows found in sheet '{sheet_name}'")
                # Still create Step 3 file but with no additional data
                data_rows = []
            
            # Apply column mapping to each data row
            mapped_data_rows = []
            for row_data in data_rows:
                mapped_row = self.apply_column_mapping(row_data)
                if mapped_row:  # Only add if mapping was successful
                    mapped_data_rows.append(mapped_row)
            
            # Populate template and create Step 3 file
            step3_file_path = self.populate_step2_template_with_data(step2_file_path, mapped_data_rows)
            
            return step3_file_path
            
        except Exception as e:
            logging.error(f"Error processing sheet '{sheet_name}' to Step 3: {str(e)}")
            return None
    
    def process_multiple_sheets_to_step3(self, uploaded_file: UploadedFile, step2_files: list) -> Dict[str, Any]:
        """
        Process multiple Step 2 files to Step 3
        
        Args:
            uploaded_file: Original uploaded Excel file
            step2_files: List of dictionaries with Step 2 file information
            
        Returns:
            Dictionary with processing results
        """
        results = {
            'success_count': 0,
            'failed_count': 0,
            'step3_files': [],
            'failed_sheets': [],
            'total_sheets': len(step2_files)
        }
        
        for step2_info in step2_files:
            sheet_name = step2_info['sheet_name']
            step2_file_path = step2_info['file_path']
            
            try:
                step3_file_path = self.process_sheet_to_step3(uploaded_file, sheet_name, step2_file_path)
                
                if step3_file_path:
                    results['success_count'] += 1
                    results['step3_files'].append({
                        'sheet_name': sheet_name,
                        'filename': Path(step3_file_path).name,
                        'file_path': step3_file_path,
                        'step2_file': step2_file_path
                    })
                else:
                    results['failed_count'] += 1
                    results['failed_sheets'].append(sheet_name)
                    
            except Exception as e:
                logging.error(f"Failed to process sheet '{sheet_name}' to Step 3: {str(e)}")
                results['failed_count'] += 1
                results['failed_sheets'].append(sheet_name)
        
        return results